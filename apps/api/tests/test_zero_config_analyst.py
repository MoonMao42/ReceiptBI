"""Acceptance-oriented tests for the zero-config analyst vertical slice."""

from __future__ import annotations

import asyncio
import base64
import contextlib
import os
import sqlite3
from pathlib import Path
from types import SimpleNamespace
from uuid import UUID, uuid4

import duckdb
import pandas as pd
import pytest
from httpx import AsyncClient
from openpyxl import Workbook
from pydantic_ai import ModelMessage, ModelResponse, ToolCallPart
from pydantic_ai.models.function import AgentInfo, FunctionModel
from pydantic_ai.models.test import TestModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.db.tables import (
    AnalysisRun,
    ArtifactRecord,
    Conversation,
    Message,
    Project,
    SemanticEntry,
)
from app.models import SSEEvent, SSEEventType
from app.services import analyst_runtime
from app.services.analysis_checkpoint import (
    CheckpointError,
    _reserve_checkpoint_revision,
    checkpoint_manifest_is_readable,
    load_runtime_checkpoint,
    recover_interrupted_analysis_runs,
    save_runtime_checkpoint,
    source_fingerprint_map,
    stable_payload_hash,
    validate_source_fingerprints,
)
from app.services.analyst_runtime import (
    PydanticAnalystRuntime,
    _aggregate_result_rows,
    _build_result_chart_code,
    _join_result_rows,
    _matching_candidate_relationship,
    _query_project_file_rows,
    _relationship_profile,
    _relationship_proof_scope,
    _render_result_chart,
    _required_trial_relationship,
    _result_profile,
    _validate_read_only,
)
from app.services.data_preflight import fingerprint_file, run_preflight
from app.services.dependency_manager import ProjectDependencyManager
from app.services.execution import ExecutionService
from app.services.golden_regression import (
    build_golden_contract,
    evaluate_golden_contract,
    find_matching_contract,
)
from app.services.project_context import ProjectRuntimeContext, load_project_context
from app.services.python_sandbox import PythonSandbox
from app.services.result_filters import apply_value_filter
from app.services.semantic_adapter import SemanticEngineAdapter, build_manifest


def _write_messy_orders(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "线上订单"
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "7 月线上订单明细"
    sheet.append([])
    sheet.append(["订单号", "门店ID", "商品品类", "实付金额", "退款状态", "日期"])
    sheet.append(["O-001", "S-01", "A类", "¥32.00", "否", "2026/07/01"])
    sheet.append(["O-001", "S-01", "A类", "¥32.00", "否", "2026/07/01"])
    sheet.append(["O-002", "S-02", "B类", "28", "已退款", "2026-07-02"])
    sheet.append(["合计", None, None, "¥92.00", None, None])
    workbook.save(path)


def test_messy_excel_preflight_preserves_original_and_creates_working_copy(tmp_path: Path):
    source = tmp_path / "orders.xlsx"
    _write_messy_orders(source)
    original_fingerprint = fingerprint_file(source)

    result = run_preflight(source, tmp_path / "working")

    assert fingerprint_file(source) == original_fingerprint
    assert result.working_path is not None and result.working_path.exists()
    assert result.status == "needs_confirmation"
    assert result.source_snapshot["ready_rows"] == 2
    assert result.source_snapshot["summary_code"] == "file_preflight"
    assert result.source_snapshot["summary_facts"] == {
        "rows": 2,
        "columns": 6,
        "automatic_issue_count": 3,
        "ambiguity_count": 1,
        "recipe_step_count": 0,
        "recipe_drift_count": 0,
    }
    assert {issue["code"] for issue in result.issues} >= {
        "header_offset",
        "duplicate_rows",
        "summary_rows",
    }
    assert result.ambiguities[0]["key"] == "revenue_refund_policy"
    assert result.ambiguities[0]["presentation_code"] == ("preflight.revenue_refund_policy")
    assert result.ambiguities[0]["option_codes"]["扣除退款"] == "exclude_refunds"
    cleaned = pd.read_parquet(result.working_path)
    assert cleaned["实付金额"].tolist() == [32.0, 28.0]
    assert cleaned["日期"].dt.strftime("%Y-%m-%d").tolist() == ["2026-07-01", "2026-07-02"]


@pytest.mark.parametrize(
    "confirmation",
    [
        {
            "key": "refund_policy",
            "question": "退款是否扣除？",
            "options": [],
            "reason": "会改变收入结论",
        },
        {
            "key": "refund_policy",
            "question": "退款是否扣除？",
            "options": ["扣除退款", " 扣除退款 "],
            "reason": "会改变收入结论",
        },
        {
            "key": "   ",
            "question": "退款是否扣除？",
            "options": ["扣除退款", "保留退款"],
            "reason": "会改变收入结论",
        },
    ],
)
def test_confirmation_contract_rejects_unusable_questions(confirmation: dict):
    with pytest.raises(ValueError):
        analyst_runtime.AnalysisReport(
            status="waiting_confirmation",
            title="确认收入口径",
            summary="需要一个业务决定",
            confirmation=confirmation,
        )


@pytest.mark.parametrize(
    ("key", "question", "reason", "expected_key"),
    [
        (
            "refund_handling",
            "计算收入时，退款订单需要扣除吗？",
            "会改变营收结论。",
            "revenue_refund_policy",
        ),
        (
            "refund_question",
            "退款金额要不要计入营收？",
            "两种算法会得到不同收入。",
            "revenue_refund_policy",
        ),
        (
            "refund_window_policy",
            "退款申请允许在几天内提交？",
            "这只会改变客服处理时限。",
            "refund_window_policy",
        ),
    ],
)
def test_confirmation_contract_emits_system_owned_decision_slots(
    key: str,
    question: str,
    reason: str,
    expected_key: str,
):
    report = analyst_runtime.AnalysisReport(
        status="waiting_confirmation",
        title="确认业务口径",
        summary="需要一个业务决定",
        confirmation={
            "key": key,
            "question": question,
            "options": ["选项一", "选项二"],
            "reason": reason,
        },
    )

    assert report.model_dump()["confirmation"]["key"] == expected_key


def test_preflight_selects_clear_data_sheet_without_interrupting_for_notes(tmp_path: Path):
    source = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "说明"
    workbook.active["A1"] = "月度经营数据"
    orders = workbook.create_sheet("订单明细")
    orders.append(["订单号", "门店ID", "金额"])
    orders.append(["O-1", "S-1", 10])
    orders.append(["O-2", "S-2", 20])
    workbook.save(source)

    result = run_preflight(source, tmp_path / "working")

    assert result.source_snapshot["reader"]["selected_sheet"] == "订单明细"
    assert not any(item["key"] == "excel_sheet_selection" for item in result.ambiguities)


def test_preflight_only_asks_when_multiple_sheets_are_plausible_tables(tmp_path: Path):
    source = tmp_path / "two-data-sheets.xlsx"
    workbook = Workbook()
    orders = workbook.active
    orders.title = "订单明细"
    orders.append(["订单号", "门店ID", "金额"])
    orders.append(["O-1", "S-1", 10])
    orders.append(["O-2", "S-2", 20])
    refunds = workbook.create_sheet("退款明细")
    refunds.append(["退款单号", "订单号", "退款金额"])
    refunds.append(["R-1", "O-1", 10])
    refunds.append(["R-2", "O-2", 20])
    workbook.save(source)

    result = run_preflight(source, tmp_path / "working")

    sheet_question = next(
        item for item in result.ambiguities if item["key"] == "excel_sheet_selection"
    )
    assert set(sheet_question["options"]) == {"退款明细", "订单明细"}


def test_recipe_replay_forces_currency_rule_without_name_hint(tmp_path: Path):
    source = tmp_path / "orders.csv"
    source.write_text("订单号,付款值\nO-1,¥32.00\nO-2,28\n", encoding="utf-8")

    baseline = run_preflight(source, tmp_path / "baseline")
    assert not pd.api.types.is_numeric_dtype(pd.read_parquet(baseline.working_path)["付款值"])

    replayed = run_preflight(
        source,
        tmp_path / "replayed",
        recipe_operations=[{"operation": "normalize_currency", "column": "付款值"}],
    )

    assert pd.read_parquet(replayed.working_path)["付款值"].tolist() == [32.0, 28.0]
    assert replayed.source_snapshot["recipe_replay"]["drift"] == []
    assert any(issue["code"] == "recipe_replayed" for issue in replayed.issues)
    assert any(
        operation["operation"] == "normalize_currency" and operation.get("replayed")
        for operation in replayed.operations
    )


def test_recipe_replay_preserves_selected_excel_sheet(tmp_path: Path):
    source = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "订单明细"
    workbook.active.append(["订单号", "付款值"])
    workbook.active.append(["O-1", "¥32"])
    summary = workbook.create_sheet("自动摘要")
    summary.append(["月份", "区域", "负责人", "说明"])
    summary.append(["七月", "华东", "小王", "经营正常"])
    workbook.save(source)

    replayed = run_preflight(
        source,
        tmp_path / "replayed",
        recipe_operations=[
            {"operation": "select_sheet", "sheet": "订单明细"},
            {"operation": "normalize_currency", "column": "付款值"},
        ],
    )

    assert replayed.source_snapshot["reader"]["selected_sheet"] == "订单明细"
    assert pd.read_parquet(replayed.working_path).columns.tolist() == ["订单号", "付款值"]
    assert replayed.source_snapshot["recipe_replay"]["drift"] == []


def test_recipe_replay_reports_missing_target_column(tmp_path: Path):
    source = tmp_path / "orders.csv"
    source.write_text("订单号,实付金额\nO-1,32\n", encoding="utf-8")

    replayed = run_preflight(
        source,
        tmp_path / "replayed",
        recipe_operations=[{"operation": "normalize_currency", "column": "付款值"}],
    )

    assert replayed.status == "needs_confirmation"
    assert replayed.source_snapshot["recipe_replay"]["drift"]
    assert any(issue["code"] == "recipe_replay_drift" for issue in replayed.issues)
    assert not any(issue["code"] == "recipe_replayed" for issue in replayed.issues)


def test_preflight_reads_common_chinese_csv_and_flags_conflicting_order_ids(tmp_path: Path):
    source = tmp_path / "orders-gb18030.csv"
    source.write_bytes(
        (
            "订单号,门店ID,金额,日期\n"
            "O-1,S-1,10,2026/07/01\n"
            "O-1,S-1,12,不是日期\n"
            "O-2,S-2,20,2026/07/03\n"
            "O-3,S-3,30,2026/07/04\n"
            "O-4,S-4,40,2026/07/05\n"
        ).encode("gb18030")
    )

    result = run_preflight(source, tmp_path / "working")
    issue_codes = {issue["code"] for issue in result.issues}

    assert result.source_snapshot["reader"]["encoding"] == "gb18030"
    assert "duplicate_business_keys" in issue_codes
    assert "invalid_date_values" in issue_codes


def test_wren_core_adapter_compiles_profile_and_plans_modeled_sql(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    context = ProjectRuntimeContext(
        project_id=uuid4(),
        name="经营分析",
        sources=[
            {
                "id": "orders",
                "name": "线上订单",
                "kind": "file",
                "format": "parquet",
                "status": "ready",
                "view_name": "online_orders",
                "profile": {
                    "summary": "线上订单明细",
                    "schema": {
                        "columns": [
                            {"name": "store_id", "dtype": "object"},
                            {"name": "category", "dtype": "object"},
                            {"name": "paid_amount", "dtype": "float64"},
                        ]
                    },
                },
            },
            {
                "id": "stores-db",
                "name": "门店库",
                "kind": "connection",
                "format": "sqlite",
                "status": "ready",
                "profile": {
                    "tables": [
                        {
                            "name": "stores",
                            "columns": [
                                {"name": "store_id", "type": "TEXT"},
                                {"name": "online_store", "type": "TEXT"},
                            ],
                        }
                    ]
                },
            },
        ],
    )

    manifest = build_manifest(context)
    assert manifest["models"][0]["columns"][2]["type"] == "DOUBLE"
    adapter = SemanticEngineAdapter(context)
    planned = adapter.transform_sql(
        "SELECT category, SUM(paid_amount) AS revenue FROM online_orders GROUP BY category",
        source_id="files",
    )
    planned_stores = adapter.transform_sql(
        "SELECT store_id, online_store FROM stores",
        source_id="stores-db",
    )

    assert adapter.status == "wren-core"
    assert adapter.compiled_backends == ["files", "stores-db"]
    assert "online_orders" in planned
    assert "stores" in planned_stores
    assert (context.project_dir / "target" / "mdl.json").exists()
    assert (context.project_dir / "target" / "mdl" / "stores_db.json").exists()


def _same_backend_relationship_context() -> ProjectRuntimeContext:
    source = {
        "id": "warehouse",
        "name": "经营库",
        "kind": "connection",
        "format": "sqlite",
        "status": "ready",
        "profile": {
            "tables": [
                {
                    "name": "orders",
                    "columns": [
                        {"name": "order_id", "type": "TEXT"},
                        {"name": "store_id", "type": "TEXT"},
                    ],
                },
                {
                    "name": "stores",
                    "columns": [
                        {"name": "store_id", "type": "TEXT"},
                        {"name": "online_store", "type": "TEXT"},
                    ],
                },
            ]
        },
    }
    definition = {
        "version": 1,
        "left": {
            "source_logical_name": "经营库",
            "source_kind": "connection",
            "table_or_view": "orders",
            "column": "store_id",
            "data_type": "TEXT",
            "schema_signature": "a" * 64,
        },
        "right": {
            "source_logical_name": "经营库",
            "source_kind": "connection",
            "table_or_view": "stores",
            "column": "store_id",
            "data_type": "TEXT",
            "schema_signature": "b" * 64,
        },
        "normalization": "exact",
        "cardinality": "many_to_one",
        "default_join": "left",
        "minimum_left_match_rate": 0.8,
        "maximum_expansion_ratio": 1.2,
    }
    return ProjectRuntimeContext(
        project_id=uuid4(),
        name="同库关系项目",
        sources=[source],
        executable_relationships={
            "orders_stores": {
                "key": "orders_stores",
                "state": "confirmed",
                "validity": "active",
                "definition": definition,
                "resolved_sources": {
                    "left": {"source_id": "warehouse"},
                    "right": {"source_id": "warehouse"},
                },
            }
        },
    )


def test_confirmed_same_backend_relationship_is_executable_and_cannot_be_bypassed(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    context = _same_backend_relationship_context()
    manifest = build_manifest(context, context.sources)

    assert manifest["relationships"] == [
        {
            "name": "orders_stores",
            "models": ["orders", "stores"],
            "joinType": "many_to_one",
            "condition": '"orders"."store_id" = "stores"."store_id"',
        }
    ]
    orders_model = next(model for model in manifest["models"] if model["name"] == "orders")
    relationship_handle = next(
        column for column in orders_model["columns"] if column.get("relationship")
    )
    assert relationship_handle["name"] == "stores"
    assert relationship_handle["relationship"] == "orders_stores"
    assert any(
        column["name"] == "stores_online_store"
        and column.get("expression") == '"stores"."online_store"'
        for column in orders_model["columns"]
    )

    adapter = SemanticEngineAdapter(context)
    planned = adapter.transform_sql(
        "SELECT order_id, stores_online_store FROM orders",
        source_id="warehouse",
    )
    normalized_plan = planned.lower().replace('"', "").replace(" ", "").replace("\n", "")
    assert "join" in normalized_plan
    assert (
        "orders.store_id=stores.store_id" in normalized_plan
        or "stores.store_id=orders.store_id" in normalized_plan
    )
    assert adapter.compiled_relationship_keys == ["orders_stores"]
    assert adapter.internal_relationship_keys == []
    assert "orders.stores_online_store" in adapter.instructions()

    database = sqlite3.connect(tmp_path / "relationship.db")
    try:
        database.executescript(
            """
            CREATE TABLE orders (order_id TEXT, store_id TEXT);
            CREATE TABLE stores (store_id TEXT, online_store TEXT);
            INSERT INTO orders VALUES ('O-1', 'S-1');
            INSERT INTO stores VALUES ('S-1', '一店');
            """
        )
        assert database.execute(planned).fetchall() == [("O-1", "一店")]
    finally:
        database.close()

    with pytest.raises(ValueError, match="不能手写 JOIN"):
        adapter.validate_sql(
            "SELECT * FROM orders JOIN stores ON orders.order_id = stores.store_id",
            source_id="warehouse",
        )


def test_wren_relationships_fail_closed_for_unverified_cross_backend_and_name_collision(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    context = _same_backend_relationship_context()
    relationship = context.executable_relationships["orders_stores"]
    relationship["validity"] = "unverified"
    assert build_manifest(context, context.sources)["relationships"] == []

    relationship["validity"] = "active"
    relationship["resolved_sources"]["right"]["source_id"] = "another-backend"
    adapter = SemanticEngineAdapter(context)
    assert adapter.compiled_relationship_keys == []
    assert adapter.internal_relationship_keys == ["orders_stores"]
    assert any(item["kind"] == "relationship_internal" for item in adapter.diagnostics)

    collision_source = {
        "id": "warehouse",
        "name": "db",
        "kind": "connection",
        "profile": {
            "tables": [
                {"name": "sales-orders", "columns": [{"name": "id", "type": "TEXT"}]},
                {
                    "name": "sales orders",
                    "columns": [{"name": "id", "type": "TEXT"}],
                },
                {
                    "name": "facts",
                    "columns": [{"name": "sales_id", "type": "TEXT"}],
                },
            ]
        },
    }
    collision_definition = relationship["definition"] | {
        "left": relationship["definition"]["left"]
        | {"table_or_view": "facts", "column": "sales_id"},
        "right": relationship["definition"]["right"]
        | {"table_or_view": "sales orders", "column": "id"},
    }
    collision_context = ProjectRuntimeContext(
        sources=[collision_source],
        executable_relationships={
            "facts_sales": {
                "key": "facts_sales",
                "validity": "active",
                "definition": collision_definition,
                "resolved_sources": {
                    "left": {"source_id": "warehouse"},
                    "right": {"source_id": "warehouse"},
                },
            }
        },
    )
    collision_manifest = build_manifest(collision_context, collision_context.sources)
    assert collision_manifest["relationships"][0]["models"] == [
        "facts",
        "db_sales_orders",
    ]


def test_wren_maps_dirty_physical_identifiers_before_relationship_planning(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    context = _same_backend_relationship_context()
    for table in context.sources[0]["profile"]["tables"]:
        for column in table["columns"]:
            if column["name"] == "store_id":
                column["name"] = "store id"
    definition = context.executable_relationships["orders_stores"]["definition"]
    definition["left"]["column"] = "store id"
    definition["right"]["column"] = "store id"

    manifest = build_manifest(context, context.sources)
    orders_model = next(model for model in manifest["models"] if model["name"] == "orders")
    mapped_column = next(
        column for column in orders_model["columns"] if column["name"] == "store_id"
    )
    assert mapped_column["expression"] == '"store id"'
    assert mapped_column["properties"]["physicalName"] == "store id"

    adapter = SemanticEngineAdapter(context)
    planned = adapter.transform_sql(
        "SELECT order_id, stores_online_store FROM orders",
        source_id="warehouse",
    )
    assert '__source."store id" AS store_id' in planned
    assert "orders.store_id = stores.store_id" in planned or (
        "stores.store_id = orders.store_id" in planned
    )
    assert "orders.store_id 对应物理字段 store id" in adapter.instructions()


@pytest.mark.asyncio
async def test_python_runtime_can_render_store_order_chart_from_rows():
    sandbox = PythonSandbox(language="zh")
    rows = [
        {"store": "一店", "orders": 12},
        {"store": "二店", "orders": 5},
    ]
    code = """
import seaborn as sns
sns.barplot(data=store_order_summary, x='store', y='orders')
print('chart-ready')
"""
    try:
        output, images = await sandbox.execute(code, sql_data={"store_order_summary": rows})
    finally:
        sandbox.cleanup()

    assert output and "chart-ready" in output
    assert len(images) == 1
    assert len(images[0]) > 1000


@pytest.mark.asyncio
async def test_generated_chart_code_uses_python_none_for_optional_color():
    rows = [{"product": "商品A", "total_profit": 338}]
    code = _build_result_chart_code(
        rows,
        chart_type="bar",
        x="product",
        y="total_profit",
        value=None,
        color=None,
        title="产品利润",
    )

    assert "hue=None" in code
    assert "hue=null" not in code
    sandbox = PythonSandbox(language="zh")
    try:
        _output, images = await sandbox.execute(code, sql_data={"chart_source": rows})
    finally:
        sandbox.cleanup()
    assert len(images) == 1


@pytest.mark.asyncio
async def test_python_runtime_exposes_all_retained_results_through_dfs_map():
    sandbox = PythonSandbox(language="zh")
    try:
        output, _images = await sandbox.execute(
            "print(int(dfs['orders']['amount'].sum() + dfs['stores']['weight'].sum()))",
            sql_data={
                "orders": [{"amount": 7}, {"amount": 5}],
                "stores": [{"weight": 3}],
            },
        )
    finally:
        sandbox.cleanup()

    assert output and output.strip() == "15"


@pytest.mark.asyncio
async def test_python_sandbox_runs_in_killable_stateful_worker():
    sandbox = PythonSandbox(language="zh")
    try:
        await sandbox.execute("intermediate_total = 41")
        output, _ = await sandbox.execute("print(intermediate_total + 1)")
        assert output and output.strip() == "42"
        assert sandbox._process is not None
        assert sandbox._process.pid != os.getpid()
    finally:
        sandbox.cleanup()

    assert sandbox._process is None


@pytest.mark.asyncio
async def test_python_sandbox_timeout_terminates_runaway_worker():
    sandbox = PythonSandbox(language="zh")
    try:
        with pytest.raises(RuntimeError, match="timeout"):
            await sandbox.execute("while True:\n    pass", timeout=1)
        assert sandbox._process is None

        output, _ = await sandbox.execute("print('recovered')")
        assert output and "recovered" in output
    finally:
        sandbox.cleanup()


def test_only_single_read_only_queries_are_allowed():
    _validate_read_only("WITH totals AS (SELECT 1 AS amount) SELECT * FROM totals")
    with pytest.raises(ValueError):
        _validate_read_only("SELECT 1; DROP TABLE orders")
    with pytest.raises(ValueError):
        _validate_read_only("SELECT pg_read_file('/etc/passwd')")
    with pytest.raises(ValueError):
        _validate_read_only("SELECT * INTO copied_orders FROM orders")


def test_result_profile_checks_keys_duplicates_nulls_and_totals():
    profile = _result_profile(
        [
            {"store_id": "S-1", "revenue": 10},
            {"store_id": "S-1", "revenue": 10},
            {"store_id": None, "revenue": 5},
        ],
        key_columns=["store_id"],
        numeric_columns=["revenue"],
    )

    assert profile["duplicate_rows"] == 1
    assert profile["keys"]["store_id"] == {
        "missing": 1,
        "unique": 1,
        "duplicate_values": 1,
    }
    assert profile["numeric"]["revenue"]["sum"] == 25


def test_relationship_profile_reports_coverage_and_cardinality():
    profile = _relationship_profile(
        [
            {"store_id": "S-1"},
            {"store_id": "S-1"},
            {"store_id": "S-2"},
            {"store_id": "S-X"},
        ],
        "store_id",
        [{"id": "S-1"}, {"id": "S-2"}, {"id": "S-3"}],
        "id",
    )

    assert profile["cardinality"] == "many_to_one"
    assert profile["left_match_rate"] == 0.75
    assert profile["right_match_rate"] == pytest.approx(2 / 3, abs=1e-6)
    assert profile["left_unmatched_examples"] == ["S-X"]


def test_relationship_profile_separates_non_null_matches_from_all_record_coverage():
    profile = _relationship_profile(
        [{"store_id": "S-1"}, {"store_id": None}, {"store_id": ""}],
        "store_id",
        [{"id": "S-1"}],
        "id",
    )

    assert profile["left_non_null_match_rate"] == 1
    assert profile["left_full_record_coverage"] == pytest.approx(1 / 3, abs=1e-6)
    assert profile["left_match_rate"] == pytest.approx(1 / 3, abs=1e-6)
    assert profile["left_null_keys"] == 2


def test_typed_cross_source_join_normalizes_ids_and_preserves_relationship_evidence():
    rows, profile = _join_result_rows(
        [
            {"门店ID": " S-01 ", "商品": "A类", "订单号": "O-1"},
            {"门店ID": "s02", "商品": "B类", "订单号": "O-2"},
            {"门店ID": "S-XX", "商品": "A类", "订单号": "O-3"},
        ],
        "门店ID",
        [
            {"store_id": "s01", "门店": "一店"},
            {"store_id": "S-02", "门店": "二店"},
        ],
        "store_id",
        normalization="auto",
    )

    assert profile["normalization"] == "identifier"
    assert profile["exact_left_match_rate"] == 0
    assert profile["left_match_rate"] == pytest.approx(2 / 3, abs=1e-6)
    assert profile["cardinality"] == "one_to_one"
    assert [row.get("门店") for row in rows] == ["一店", "二店", None]


def test_typed_aggregation_creates_a_validatable_store_summary():
    rows = _aggregate_result_rows(
        [
            {"门店": "一店", "订单号": "O-1"},
            {"门店": "一店", "订单号": "O-2"},
            {"门店": "二店", "订单号": "O-3"},
        ],
        group_by=["门店"],
        operation="nunique",
        value_column="订单号",
        output_column="订单数",
    )

    assert rows == [
        {"门店": "一店", "订单数": 2},
        {"门店": "二店", "订单数": 1},
    ]


@pytest.mark.asyncio
async def test_typed_chart_renderer_creates_bar_chart_without_model_written_code():
    sandbox = PythonSandbox(language="zh")
    try:
        _, images, code = await _render_result_chart(
            sandbox,
            [
                {"门店": "一店", "订单数": 12},
                {"门店": "二店", "订单数": 5},
            ],
            chart_type="bar",
            x="门店",
            y="订单数",
            value=None,
            color=None,
            title="门店订单量比较",
        )
    finally:
        sandbox.cleanup()

    assert "barplot" in code
    assert len(images) == 1
    assert len(images[0]) > 1000


def test_project_file_query_can_only_read_materialized_project_sources(tmp_path: Path):
    allowed = tmp_path / "allowed.parquet"
    secret = tmp_path / "secret.csv"
    pd.DataFrame([{"store_id": "S-1", "revenue": 10}]).to_parquet(allowed, index=False)
    secret.write_text("token\nshould-not-be-readable\n", encoding="utf-8")
    sources = [
        {
            "kind": "file",
            "working_uri": str(allowed),
            "view_name": "orders",
        }
    ]

    rows, truncated, available = _query_project_file_rows(
        sources,
        "SELECT * FROM orders",
        tmp_path / "project",
    )

    assert rows == [{"store_id": "S-1", "revenue": 10}]
    assert truncated is False
    assert available == ["orders"]

    escaped_secret = str(secret).replace("'", "''")
    with pytest.raises(duckdb.PermissionException, match="file system operations are disabled"):
        _query_project_file_rows(
            sources,
            f"SELECT * FROM read_csv_auto('{escaped_secret}')",
            tmp_path / "project",
        )


def test_project_dependency_manifest_tracks_requested_and_resolved_versions(tmp_path: Path):
    manager = ProjectDependencyManager(tmp_path / "project")
    metadata_dir = manager.target / "demo_pkg-1.2.3.dist-info"
    metadata_dir.mkdir(parents=True)
    (metadata_dir / "METADATA").write_text(
        "Metadata-Version: 2.1\nName: demo-pkg\nVersion: 1.2.3\n",
        encoding="utf-8",
    )

    manager._record_install(["demo-pkg>=1"])

    assert manager.describe() == {
        "requested": ["demo-pkg>=1"],
        "installed": [{"name": "demo-pkg", "version": "1.2.3"}],
    }


@pytest.mark.asyncio
async def test_python_execution_auto_installs_a_missing_project_module_once(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: TestModel())
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(project_id=uuid4(), name="依赖项目"),
    )
    installed_requests: list[list[str]] = []

    async def fake_install(packages: list[str], *args, **kwargs) -> str:
        del args, kwargs
        installed_requests.append(packages)
        package_dir = runtime.deps.dependency_manager.target / packages[0]
        package_dir.mkdir(parents=True, exist_ok=True)
        (package_dir / "__init__.py").write_text("VALUE = 42\n", encoding="utf-8")
        return "installed"

    monkeypatch.setattr(runtime.deps.dependency_manager, "install", fake_install)
    code = "import receiptbi_auto_dep_probe\nprint(receiptbi_auto_dep_probe.VALUE)"
    try:
        output, images, installed = await runtime._execute_python_code(code, sql_data={})
        second_output, _, second_installed = await runtime._execute_python_code(code, sql_data={})
    finally:
        runtime.deps.python_sandbox.cleanup()

    assert output and output.strip() == "42"
    assert second_output and second_output.strip() == "42"
    assert images == []
    assert installed == ["receiptbi_auto_dep_probe"]
    assert second_installed == []
    assert installed_requests == [["receiptbi_auto_dep_probe"]]


@pytest.mark.asyncio
async def test_analysis_run_persists_report_table_chart_and_hidden_evidence(
    client: AsyncClient,
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path / "projects")
    project = Project(name="经营分析项目")
    db_session.add(project)
    await db_session.flush()
    run = AnalysisRun(project_id=project.id, query="比较各商品品类与门店的订单量")
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)
    image = base64.b64encode(b"\x89PNG\r\n\x1a\nreceiptbi").decode()

    await service._persist_project_result(
        run,
        {
            "analysis_state": "completed",
            "report": {
                "title": "门店订单量比较",
                "summary": "一店订单量最高。",
                "metrics": [{"label": "匹配率", "value": "95%"}],
                "visualization": {
                    "version": 1,
                    "type": "bar",
                    "title": "门店订单量",
                    "data_ref": {
                        "result_name": "store_order_summary",
                        "result_hash": "verified-chart-hash",
                    },
                    "encoding": {
                        "x": {"field": "store"},
                        "y": [{"field": "orders", "format": "integer"}],
                    },
                },
            },
            "data": [{"category": "A类", "store": "一店", "orders": 12}],
            "rows_count": 1,
            "python_images": [image],
            "tool_history": [
                {
                    "kind": "file_sql",
                    "sql": "SELECT category, store, COUNT(*) AS orders FROM joined GROUP BY 1, 2",
                    "result_name": "store_order_summary",
                },
                {
                    "kind": "relationship_validation",
                    "profile": {"left_match_rate": 0.95},
                },
                {
                    "kind": "validation",
                    "result_name": "store_order_summary",
                    "profile": {
                        "columns": ["category", "store", "orders"],
                        "keys": {"category": {}, "store": {}},
                        "numeric": {"orders": {"count": 1}},
                        "truncated": False,
                    },
                },
            ],
            "knowledge_proposals": [
                {
                    "key": "revenue_refund_policy",
                    "value": "收入扣除退款",
                    "state": "confirmed",
                    "confidence": 1,
                    "source": "user",
                    "evidence": [{"description": "用户确认"}],
                }
            ],
        },
    )

    result = await db_session.execute(
        select(ArtifactRecord).where(ArtifactRecord.analysis_run_id == run.id)
    )
    artifacts = list(result.scalars())
    assert {artifact.kind for artifact in artifacts} == {
        "report",
        "metric",
        "table",
        "chart",
        "evidence",
    }
    structured_chart = next(
        artifact
        for artifact in artifacts
        if artifact.kind == "chart"
        and artifact.technical_details.get("source") == "structured_visualization"
    )
    assert structured_chart.technical_details["result_name"] == "store_order_summary"
    assert structured_chart.technical_details["result_hash"] == "verified-chart-hash"
    chart = next(
        artifact
        for artifact in artifacts
        if artifact.kind == "chart" and artifact.technical_details.get("source") == "python"
    )
    response = await client.get(
        f"/api/v1/projects/{project.id}/analysis-runs/{run.id}/artifacts/{chart.id}/file"
    )
    assert response.status_code == 200
    assert response.content == base64.b64decode(image)
    await db_session.refresh(project)
    assert project.extra_data["golden_scenarios"][0]["query"] == "比较各商品品类与门店的订单量"
    knowledge_result = await db_session.execute(
        select(SemanticEntry).where(SemanticEntry.project_id == project.id)
    )
    knowledge = list(knowledge_result.scalars())
    assert any(
        item.key.startswith("verified_query:") and item.state == "confirmed" for item in knowledge
    )


@pytest.mark.asyncio
async def test_structured_chart_artifact_keeps_legacy_result_reference(
    db_session: AsyncSession,
):
    project = Project(name="旧图表兼容项目")
    db_session.add(project)
    await db_session.flush()
    run = AnalysisRun(project_id=project.id, query="查看旧版图表")
    db_session.add(run)
    await db_session.commit()

    await ExecutionService(db_session, project_id=project.id)._persist_project_result(
        run,
        {
            "analysis_state": "completed",
            "report": {
                "status": "completed",
                "title": "旧版月度销售",
                "summary": "已生成旧版图表。",
                "visualization": {
                    "type": "bar",
                    "result_name": "legacy_monthly_sales",
                    "result_hash": "legacy-chart-hash",
                },
            },
        },
    )

    result = await db_session.execute(
        select(ArtifactRecord).where(
            ArtifactRecord.analysis_run_id == run.id,
            ArtifactRecord.kind == "chart",
        )
    )
    chart = result.scalar_one()
    assert chart.technical_details == {
        "source": "structured_visualization",
        "result_name": "legacy_monthly_sales",
        "result_hash": "legacy-chart-hash",
    }


@pytest.mark.asyncio
async def test_explicit_confirmation_creates_golden_contract_without_model_proposal(
    db_session: AsyncSession,
):
    project = Project(name="确定性学习项目")
    db_session.add(project)
    await db_session.flush()
    db_session.add(
        SemanticEntry(
            project_id=project.id,
            key="revenue_refund_policy",
            value="扣除退款",
            entry_type="business_rule",
            state="confirmed",
            confidence=1,
            evidence=[{"kind": "explicit_confirmation"}],
            source="user",
        )
    )
    run = AnalysisRun(project_id=project.id, query="我选择扣除退款")
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)
    _, rule_evidence = apply_value_filter(
        [
            {"退款状态": "否", "category": "A类", "store": "一店", "orders": 12},
            {"退款状态": "已退款", "category": "B类", "store": "二店", "orders": 1},
        ],
        rule_key="revenue_refund_policy",
        rule_value="扣除退款",
        column="退款状态",
        operator="exclude",
        values=["已退款"],
    )

    await service._persist_project_result(
        run,
        {
            "analysis_state": "completed",
            "report": {"title": "门店订单量比较", "summary": "已按确认口径完成"},
            "data": [{"category": "A类", "store": "一店", "orders": 12}],
            "tool_history": [
                rule_evidence,
                {
                    "kind": "file_sql",
                    "sql": "SELECT category, store, COUNT(*) AS orders FROM joined GROUP BY 1, 2",
                    "result_name": "store_order_summary",
                },
                {
                    "kind": "validation",
                    "result_name": "store_order_summary",
                    "profile": {
                        "columns": ["category", "store", "orders"],
                        "keys": {"category": {}, "store": {}},
                        "numeric": {"orders": {"count": 1}},
                        "truncated": False,
                    },
                },
            ],
            "knowledge_proposals": [],
            "confirmed_corrections": [
                {
                    "key": "revenue_refund_policy",
                    "value": "扣除退款",
                    "applied": True,
                    "conflict": False,
                    "task_query": "比较各商品品类与门店的订单量",
                }
            ],
        },
    )

    await db_session.refresh(project)
    contract = project.extra_data["golden_scenarios"][0]
    assert contract["version"] == 1
    assert contract["query"] == "比较各商品品类与门店的订单量"
    assert contract["result"]["required_columns"] == ["category", "orders", "store"]
    result = await db_session.execute(
        select(SemanticEntry).where(
            SemanticEntry.project_id == project.id,
            SemanticEntry.entry_type == "verified_query",
        )
    )
    assert result.scalar_one().state == "confirmed"


@pytest.mark.asyncio
async def test_refund_confirmation_without_execution_evidence_is_not_learned_as_golden(
    db_session: AsyncSession,
):
    project = Project(name="拒绝伪学习项目")
    db_session.add(project)
    await db_session.flush()
    db_session.add(
        SemanticEntry(
            project_id=project.id,
            key="revenue_refund_policy",
            value="扣除退款",
            entry_type="business_rule",
            state="confirmed",
            confidence=1,
            evidence=[{"kind": "explicit_confirmation"}],
            source="user",
        )
    )
    run = AnalysisRun(project_id=project.id, query="扣除退款后比较各商品品类与门店的订单量")
    db_session.add(run)
    await db_session.commit()

    await ExecutionService(db_session, project_id=project.id)._persist_project_result(
        run,
        {
            "analysis_state": "completed",
            "report": {"title": "缺少口径证据", "summary": "不应学习"},
            "data": [{"category": "B类", "store": "二店", "orders": 1}],
            "tool_history": [
                {
                    "kind": "file_sql",
                    "sql": "SELECT category, store, COUNT(*) AS orders FROM orders GROUP BY 1, 2",
                    "result_name": "store_order_summary",
                },
                {
                    "kind": "validation",
                    "result_name": "store_order_summary",
                    "profile": {
                        "columns": ["category", "store", "orders"],
                        "numeric": {"orders": {"count": 1}},
                        "truncated": False,
                    },
                },
            ],
            "knowledge_proposals": [],
            "confirmed_corrections": [
                {
                    "key": "revenue_refund_policy",
                    "value": "扣除退款",
                    "applied": True,
                    "conflict": False,
                    "task_query": "比较各商品品类与门店的订单量",
                }
            ],
        },
    )

    await db_session.refresh(project)
    assert (project.extra_data or {}).get("golden_scenarios") is None
    verified_result = await db_session.execute(
        select(SemanticEntry).where(
            SemanticEntry.project_id == project.id,
            SemanticEntry.entry_type == "verified_query",
        )
    )
    assert verified_result.scalar_one().state == "candidate"


def test_new_month_values_pass_when_golden_structure_and_relationship_hold():
    baseline_history = [
        {
            "kind": "join",
            "profile": {
                "left_key": "store_id",
                "right_key": "store_id",
                "normalization": "identifier",
                "left_match_rate": 0.95,
                "cardinality": "many_to_one",
                "expansion_ratio": 1,
            },
        },
        {
            "kind": "validation",
            "profile": {
                "columns": ["category", "store", "orders"],
                "keys": {"category": {}, "store": {}},
                "numeric": {"orders": {"count": 1}},
                "truncated": False,
            },
        },
    ]
    knowledge = [{"key": "revenue_refund_policy", "value": "扣除退款"}]
    contract = build_golden_contract(
        query="比较各商品品类与门店的订单量",
        confirmed_knowledge=knowledge,
        sources=[{"name": "订单", "fingerprint": "july"}],
        tool_history=baseline_history,
        result_rows=[{"category": "A类", "store": "一店", "orders": 12}],
    )
    assert contract is not None
    august_history = [
        {
            "kind": "join",
            "profile": {
                "left_key": "store_id",
                "right_key": "store_id",
                "normalization": "identifier",
                "left_match_rate": 0.92,
                "cardinality": "many_to_one",
                "expansion_ratio": 1,
            },
        },
        baseline_history[-1],
    ]

    failures = evaluate_golden_contract(
        contract,
        confirmed_knowledge=knowledge,
        sources=[{"name": "订单", "fingerprint": "august"}],
        tool_history=august_history,
        result_rows=[{"category": "A类", "store": "一店", "orders": 27}],
    )

    assert failures == []


def test_candidate_relationship_identity_is_bound_into_golden_regression() -> None:
    candidate = {
        "key": "relationship:orders:stores",
        "definition_hash": "a" * 64,
        "definition": {
            "left": {"column": "门店ID"},
            "right": {"column": "store_id"},
        },
        "resolved_sources": {
            "left": {"source_id": "orders-source", "table_or_view": "orders"},
            "right": {"source_id": "stores-source", "table_or_view": "stores"},
        },
    }
    matched = _matching_candidate_relationship(
        [candidate],
        left_endpoints={("orders-source", "orders")},
        right_endpoints={("stores-source", "stores")},
        left_key="门店ID",
        right_key="store_id",
    )
    assert matched is candidate
    assert (
        _matching_candidate_relationship(
            [candidate],
            left_endpoints={("orders-source", "order_summary")},
            right_endpoints={("stores-source", "stores")},
            left_key="门店ID",
            right_key="store_id",
        )
        is None
    )

    full_metadata = {
        "query_scope": "full",
        "result_completeness": "complete",
        "source_refs": [
            {
                "source_id": "orders-source",
                "table_or_view": "orders",
                "query_scope": "full",
            }
        ],
    }
    other_full_metadata = {
        **full_metadata,
        "source_refs": [
            {
                "source_id": "stores-source",
                "table_or_view": "stores",
                "query_scope": "full",
            }
        ],
    }
    filtered_metadata = {
        **full_metadata,
        "query_scope": "filtered",
        "source_refs": [
            {
                **full_metadata["source_refs"][0],
                "query_scope": "filtered",
            }
        ],
    }
    assert _relationship_proof_scope(full_metadata, other_full_metadata) == {
        "evidence_origin": "system",
        "evidence_scope": "full_relation",
        "completeness": "complete",
        "reusable_proof_eligible": True,
    }
    assert (
        _relationship_proof_scope(filtered_metadata, other_full_metadata)["evidence_scope"]
        == "current_result"
    )
    top_level_full_but_filtered_ref = {
        **full_metadata,
        "source_refs": [{**full_metadata["source_refs"][0], "query_scope": "filtered"}],
    }
    assert (
        _relationship_proof_scope(top_level_full_but_filtered_ref, other_full_metadata)[
            "reusable_proof_eligible"
        ]
        is False
    )
    missing_ref_scope = {
        **full_metadata,
        "source_refs": [
            {
                "source_id": "orders-source",
                "table_or_view": "orders",
            }
        ],
    }
    assert (
        _relationship_proof_scope(missing_ref_scope, other_full_metadata)["reusable_proof_eligible"]
        is False
    )
    assert (
        _relationship_proof_scope(full_metadata, full_metadata)["reusable_proof_eligible"] is False
    )
    assert _relationship_proof_scope(
        full_metadata,
        other_full_metadata,
        output_truncated=True,
    ) == {
        "evidence_origin": "system",
        "evidence_scope": "current_result",
        "completeness": "partial",
        "reusable_proof_eligible": False,
    }
    missing_table_metadata = {
        **full_metadata,
        "source_refs": [{"source_id": "orders-source", "query_scope": "full"}],
    }
    assert _relationship_proof_scope(missing_table_metadata, other_full_metadata) == {
        "evidence_origin": "system",
        "evidence_scope": "current_result",
        "completeness": "partial",
        "reusable_proof_eligible": False,
    }

    source_refs = [
        {"source_logical_name": "orders", "source_kind": "file"},
        {"source_logical_name": "stores", "source_kind": "connection"},
    ]
    relationship = {
        "kind": "join",
        "candidate_relationship_key": candidate["key"],
        "definition_hash": candidate["definition_hash"],
        "source_refs": source_refs,
        "profile": {
            "left_key": "门店ID",
            "right_key": "store_id",
            "normalization": "identifier",
            "left_match_rate": 1,
            "expansion_ratio": 1,
        },
    }
    validation = {
        "kind": "validation",
        "profile": {
            "columns": ["category", "store", "orders"],
            "keys": {"category": {}, "store": {}},
            "numeric": {"orders": {"sum": 5}},
            "truncated": False,
        },
    }
    contract = build_golden_contract(
        query="比较各商品品类与门店的订单量",
        confirmed_knowledge=[],
        sources=[],
        tool_history=[relationship, validation],
        result_rows=[{"category": "A类", "store": "一店", "orders": 5}],
    )
    assert contract is not None
    assert contract["relationships"][0]["relationship_key"] == candidate["key"]
    assert contract["relationships"][0]["definition_hash"] == candidate["definition_hash"]

    assert (
        evaluate_golden_contract(
            contract,
            confirmed_knowledge=[],
            sources=[],
            tool_history=[relationship, validation],
            result_rows=[{"category": "A类", "store": "一店", "orders": 8}],
        )
        == []
    )
    failures = evaluate_golden_contract(
        contract,
        confirmed_knowledge=[],
        sources=[],
        tool_history=[
            {**relationship, "candidate_relationship_key": None},
            validation,
        ],
        result_rows=[{"category": "A类", "store": "一店", "orders": 8}],
    )
    assert any("缺少关联验证" in failure for failure in failures)


def test_only_the_current_required_correction_can_select_a_candidate_relationship() -> None:
    candidate = {
        "id": "semantic-relationship",
        "active_revision_id": "semantic-revision",
        "key": "relationship:orders:stores",
        "state": "candidate",
        "validity": "unverified",
        "execution_state": "needs_validation",
        "definition_hash": "d" * 64,
        "definition": {"left": {}, "right": {}},
    }
    context = ProjectRuntimeContext(
        candidate_relationships=[candidate],
        required_correction={
            "id": "correction-1",
            "correction_type": "relationship_rule",
            "target_key": candidate["key"],
            "semantic_entry_id": candidate["id"],
            "expected_active_revision_id": candidate["active_revision_id"],
            "definition_hash": candidate["definition_hash"],
            "execution_state": "needs_validation",
            "executable": True,
        },
    )

    assert _required_trial_relationship(context, candidate["key"]) is candidate
    assert _required_trial_relationship(context, "relationship:other") is None
    original = dict(context.required_correction or {})
    for field, value in (
        ("semantic_entry_id", "semantic-other"),
        ("expected_active_revision_id", "revision-other"),
        ("definition_hash", "e" * 64),
        ("target_key", "relationship:other"),
    ):
        context.required_correction = {**original, field: value}
        assert _required_trial_relationship(context, candidate["key"]) is None
    context.required_correction = None
    assert _required_trial_relationship(context, candidate["key"]) is None


def test_golden_contract_rejects_missing_result_column_and_low_join_coverage():
    history = [
        {
            "kind": "join",
            "profile": {
                "left_key": "store_id",
                "right_key": "store_id",
                "left_match_rate": 0.95,
                "expansion_ratio": 1,
            },
        },
        {
            "kind": "validation",
            "profile": {"columns": ["category", "store", "orders"], "truncated": False},
        },
    ]
    contract = build_golden_contract(
        query="比较各商品品类与门店的订单量",
        confirmed_knowledge=[],
        sources=[],
        tool_history=history,
        result_rows=[{"category": "A类", "store": "一店", "orders": 12}],
    )
    assert contract is not None
    failures = evaluate_golden_contract(
        contract,
        confirmed_knowledge=[],
        sources=[],
        tool_history=[
            {
                "kind": "join",
                "profile": {
                    "left_key": "store_id",
                    "right_key": "store_id",
                    "left_match_rate": 0.4,
                    "expansion_ratio": 1,
                },
            },
            {"kind": "validation", "profile": {"columns": ["category", "orders"]}},
        ],
        result_rows=[{"category": "A类", "orders": 12}],
    )
    assert any("缺少字段" in failure for failure in failures)
    assert any("覆盖率" in failure for failure in failures)


@pytest.mark.asyncio
async def test_pydantic_agent_loop_returns_a_typed_business_report(
    monkeypatch: pytest.MonkeyPatch,
):
    test_model = TestModel(
        call_tools=["inspect_project_data"],
        custom_output_args={
            "status": "completed",
            "title": "门店订单量分析",
            "summary": "一店的A类关联最强。",
            "findings": ["A类在一店的订单数最高"],
            "metrics": [{"label": "最高组合", "value": "A类 × 一店"}],
            "evidence": ["对门店和品类交叉汇总后比较"],
            "next_actions": [
                {
                    "kind": "deepen",
                    "label": "拆分高峰时段",
                    "prompt": "把A类在一店的订单按时段拆分，找出最强时段",
                    "reason": "当前只知道门店组合最强，还不知道由哪个时段推动",
                    "recommended": True,
                }
            ],
        },
    )
    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: test_model)
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="经营分析项目"),
    )

    events = [event async for event in runtime.execute(query="调查各门店订单量")]

    result = next(event for event in events if event.type == SSEEventType.RESULT)
    assert result.data["report"]["title"] == "门店订单量分析"
    assert result.data["analysis_state"] == "completed"
    assert result.data["report"]["follow_ups"] == ["把A类在一店的订单按时段拆分，找出最强时段"]
    progress = [event.data for event in events if event.type == SSEEventType.PROGRESS]
    assert progress[0]["stage"] == "understanding"
    assert progress[-1]["stage"] == "completed"
    assert any(item["message"] == "已检查当前项目的数据和业务口径" for item in progress)


@pytest.mark.asyncio
async def test_runtime_returns_the_validated_result_instead_of_the_last_temporary_frame():
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="校验结果项目"),
    )

    class AgentWithTemporaryFrameAfterValidation:
        async def run(self, prompt, deps):
            del prompt
            deps.dataframes["validated_summary"] = [{"门店": "一店", "订单数": 2}]
            deps.tool_history.append(
                {
                    "kind": "validation",
                    "result_name": "validated_summary",
                    "result_hash": stable_payload_hash([{"门店": "一店", "订单数": 2}]),
                    "profile": {"truncated": False},
                }
            )
            deps.validated_results.add("validated_summary")
            deps.dataframes["temporary_probe"] = [{"调试值": 999}]
            return SimpleNamespace(
                output=analyst_runtime.AnalysisReport(
                    status="completed",
                    title="已校验结果",
                    summary="最终报告使用已校验的门店汇总。",
                )
            )

    runtime.agent = AgentWithTemporaryFrameAfterValidation()

    events = [event async for event in runtime.execute(query="汇总门店订单")]
    result = next(event for event in events if event.type == SSEEventType.RESULT)

    assert result.data["result_name"] == "validated_summary"
    assert result.data["data"] == [{"门店": "一店", "订单数": 2}]
    assert result.data["rows_count"] == 1


@pytest.mark.asyncio
async def test_cancelling_runtime_also_cancels_the_inner_model_task():
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="取消调查项目"),
    )
    inner_cancelled = asyncio.Event()

    class BlockingAgent:
        async def run(self, prompt, deps):
            del prompt, deps
            try:
                await asyncio.Future()
            finally:
                inner_cancelled.set()

    runtime.agent = BlockingAgent()
    outer = asyncio.create_task(runtime._run_agent("调查销售", None))
    await asyncio.sleep(0)
    outer.cancel()
    with contextlib.suppress(asyncio.CancelledError):
        await outer

    assert inner_cancelled.is_set()


@pytest.mark.asyncio
async def test_missing_data_returns_an_explicit_product_action(
    monkeypatch: pytest.MonkeyPatch,
):
    test_model = TestModel(
        call_tools=["inspect_project_data"],
        custom_output_args={
            "status": "needs_data",
            "title": "还需要订单和门店资料",
            "summary": "目前没有足够的数据生成门店与商品品类的关联结论。",
            "action": {
                "kind": "add_data",
                "label": "补充两类资料",
                "reason": "需要订单明细与门店信息才能核对关联并比较各商品品类与门店的订单量。",
                "requested_data": [
                    "包含门店编号、商品品类和订单金额的订单明细",
                    "包含门店编号和门店名称的门店资料",
                ],
            },
        },
    )
    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: test_model)
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="经营分析项目"),
    )

    events = [event async for event in runtime.execute(query="比较各商品品类与门店的订单量")]

    result = next(event for event in events if event.type == SSEEventType.RESULT)
    assert result.data["analysis_state"] == "needs_attention"
    assert result.data["report"]["status"] == "needs_data"
    assert result.data["report"]["action"]["kind"] == "add_data"
    progress = [event.data for event in events if event.type == SSEEventType.PROGRESS]
    assert progress[0]["stage"] == "understanding"
    assert progress[-1]["stage"] == "needs_attention"
    assert any(item["message"] == "已检查当前项目的数据和业务口径" for item in progress)


@pytest.mark.asyncio
async def test_business_confirmation_requires_a_pending_question(
    monkeypatch: pytest.MonkeyPatch,
):
    test_model = TestModel(
        call_tools=["inspect_project_data"],
        custom_output_args={
            "status": "completed",
            "title": "口径已记录",
            "summary": "后续会沿用该口径。",
        },
    )
    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: test_model)
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="经营分析项目"),
    )

    events = [
        event
        async for event in runtime.execute(
            query="扣除退款",
            history=[
                {
                    "role": "assistant",
                    "content": "计算收入时，退款订单需要扣除吗？",
                    "confirmation": {
                        "key": "refund_handling",
                        "question": "计算收入时，退款订单需要扣除吗？",
                        "options": ["扣除退款", "保留退款订单"],
                    },
                }
            ],
        )
    ]

    assert any(event.type == SSEEventType.RESULT for event in events)
    assert runtime.deps.user_confirmation is True
    assert runtime.deps.pending_confirmation["key"] == "revenue_refund_policy"

    second_runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="经营分析项目"),
    )
    _ = [event async for event in second_runtime.execute(query="我确认")]
    assert second_runtime.deps.user_confirmation is False


@pytest.mark.asyncio
async def test_pending_confirmation_proposal_uses_the_pending_canonical_slot(
    monkeypatch: pytest.MonkeyPatch,
):
    calls = 0

    def model(_messages: list[ModelMessage], info: AgentInfo) -> ModelResponse:
        nonlocal calls
        calls += 1
        if calls == 1:
            return ModelResponse(
                parts=[
                    ToolCallPart(
                        "propose_project_knowledge",
                        {
                            "key": "refund_window_policy",
                            "value": "扣除退款",
                            "evidence": "用户选择了当前待确认选项",
                        },
                    )
                ]
            )
        return ModelResponse(
            parts=[
                ToolCallPart(
                    info.output_tools[0].name,
                    {
                        "status": "completed",
                        "title": "口径已记录",
                        "summary": "已按用户选择记录当前业务口径。",
                    },
                )
            ]
        )

    monkeypatch.setattr(
        analyst_runtime,
        "build_pydantic_model",
        lambda _config: FunctionModel(model),
    )
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="口径归一项目"),
    )

    events = [
        event
        async for event in runtime.execute(
            query="我选择扣除退款",
            history=[
                {
                    "role": "assistant",
                    "confirmation": {
                        "key": "refund_policy",
                        "question": "计算收入时，退款订单需要扣除吗？",
                        "options": ["扣除退款", "保留退款订单"],
                    },
                }
            ],
        )
    ]

    assert any(event.type == SSEEventType.RESULT for event in events)
    assert runtime.deps.pending_confirmation["key"] == "revenue_refund_policy"
    assert runtime.deps.knowledge_proposals == [
        {
            "key": "revenue_refund_policy",
            "value": "扣除退款",
            "entry_type": "business_rule",
            "evidence": [{"description": "用户选择了当前待确认选项"}],
            "confidence": 0.8,
            "state": "confirmed",
            "source": "user",
        }
    ]


@pytest.mark.asyncio
async def test_apply_confirmed_rule_writes_only_the_canonical_rule_key(
    monkeypatch: pytest.MonkeyPatch,
):
    definition = {
        "version": 1,
        "kind": "business_rule_strategy",
        "rule_key": "revenue_refund_policy",
        "selected_option": "扣除退款",
        "action": {
            "kind": "value_filter",
            "column": "退款状态",
            "operator": "exclude",
            "values": ["已退款"],
            "observed_values": ["否", "已退款"],
        },
    }
    calls = 0

    def model(_messages: list[ModelMessage], info: AgentInfo) -> ModelResponse:
        nonlocal calls
        calls += 1
        if calls == 1:
            return ModelResponse(
                parts=[
                    ToolCallPart(
                        "apply_confirmed_rule",
                        {
                            "source_result": "orders",
                            "rule_key": "refund_handling",
                            "result_name": "effective_orders",
                            "purpose": "按确认口径扣除退款订单",
                        },
                    )
                ]
            )
        return ModelResponse(
            parts=[
                ToolCallPart(
                    info.output_tools[0].name,
                    {
                        "status": "completed",
                        "title": "有效订单",
                        "summary": "已按确认口径处理订单。",
                    },
                )
            ]
        )

    monkeypatch.setattr(
        analyst_runtime,
        "build_pydantic_model",
        lambda _config: FunctionModel(model),
    )
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(
            name="规则归一项目",
            confirmed_knowledge=[
                {
                    "id": "semantic-refund-policy",
                    "key": "refund_policy",
                    "value": "扣除退款",
                    "state": "confirmed",
                    "execution_state": "verified",
                    "definition": definition,
                }
            ],
        ),
    )
    runtime.deps.dataframes["orders"] = [
        {"订单号": "O-1", "退款状态": "否"},
        {"订单号": "O-2", "退款状态": "已退款"},
    ]
    runtime.deps.result_metadata["orders"] = {
        "source_refs": [],
        "result_completeness": "complete",
        "truncated": False,
    }

    events = [event async for event in runtime.execute(query="按确认口径处理订单")]

    assert any(event.type == SSEEventType.RESULT for event in events)
    application = next(
        item
        for item in runtime.deps.tool_history
        if item.get("kind") == "business_rule_application"
    )
    replay = next(
        item for item in runtime.deps.replay_journal if item.get("op") == "apply_confirmed_rule"
    )
    assert application["rule_key"] == "revenue_refund_policy"
    assert replay["rule_key"] == "revenue_refund_policy"
    assert runtime.deps.result_metadata["effective_orders"]["business_rule"]["rule_key"] == (
        "revenue_refund_policy"
    )


@pytest.mark.asyncio
async def test_report_completion_compares_pending_and_evidence_by_canonical_slot(
    monkeypatch: pytest.MonkeyPatch,
):
    definition = {
        "version": 1,
        "kind": "business_rule_strategy",
        "rule_key": "revenue_refund_policy",
        "selected_option": "扣除退款",
        "action": {
            "kind": "identity",
            "column": "退款状态",
            "observed_values": ["否", "已退款"],
        },
    }

    def model(_messages: list[ModelMessage], info: AgentInfo) -> ModelResponse:
        return ModelResponse(
            parts=[
                ToolCallPart(
                    info.output_tools[0].name,
                    {
                        "status": "completed",
                        "title": "口径已应用",
                        "summary": "已按确认口径继续调查。",
                    },
                )
            ]
        )

    monkeypatch.setattr(
        analyst_runtime,
        "build_pydantic_model",
        lambda _config: FunctionModel(model),
    )
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(
            name="完成门禁归一项目",
            sources=[
                {
                    "id": "orders-source",
                    "name": "订单",
                    "kind": "file",
                    "source_uri": "/tmp/orders.parquet",
                }
            ],
            confirmed_knowledge=[
                {
                    "key": "revenue_refund_policy",
                    "value": "扣除退款",
                    "state": "confirmed",
                    "definition": definition,
                }
            ],
        ),
    )
    runtime.deps.tool_history.append(
        {"kind": "business_rule_application", "rule_key": "refund_handling"}
    )

    events = [
        event
        async for event in runtime.execute(
            query="我选择扣除退款",
            history=[
                {
                    "role": "assistant",
                    "confirmation": {
                        "key": "refund_policy",
                        "question": "计算收入时，退款订单需要扣除吗？",
                        "options": ["扣除退款", "保留退款订单"],
                    },
                }
            ],
        )
    ]

    result = next(event for event in events if event.type == SSEEventType.RESULT)
    assert result.data["analysis_state"] == "completed"
    assert runtime.deps.pending_confirmation["key"] == "revenue_refund_policy"


@pytest.mark.asyncio
async def test_report_confirmation_reuses_known_canonical_slot_instead_of_asking_again(
    monkeypatch: pytest.MonkeyPatch,
):
    calls = 0

    def model(_messages: list[ModelMessage], info: AgentInfo) -> ModelResponse:
        nonlocal calls
        calls += 1
        if calls == 1:
            return ModelResponse(
                parts=[
                    ToolCallPart(
                        info.output_tools[0].name,
                        {
                            "status": "waiting_confirmation",
                            "title": "再次确认收入口径",
                            "summary": "模型尝试重复询问已确认口径。",
                            "confirmation": {
                                "key": "refund_handling",
                                "question": "计算收入时，退款订单需要扣除吗？",
                                "options": ["扣除退款", "保留退款订单"],
                                "reason": "不同选择会改变营收结论。",
                            },
                        },
                    )
                ]
            )
        return ModelResponse(
            parts=[
                ToolCallPart(
                    info.output_tools[0].name,
                    {
                        "status": "completed",
                        "title": "已复用口径",
                        "summary": "已直接复用项目中确认的收入口径。",
                    },
                )
            ]
        )

    monkeypatch.setattr(
        analyst_runtime,
        "build_pydantic_model",
        lambda _config: FunctionModel(model),
    )
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(
            name="已确认口径项目",
            confirmed_knowledge=[
                {
                    "key": "refund_policy",
                    "value": "扣除退款",
                    "state": "confirmed",
                }
            ],
        ),
    )

    events = [event async for event in runtime.execute(query="继续分析本月收入")]

    result = next(event for event in events if event.type == SSEEventType.RESULT)
    assert calls == 2
    assert result.data["analysis_state"] == "completed"
    assert result.data["report"]["confirmation"] is None


@pytest.mark.asyncio
async def test_selected_business_option_is_persisted_without_model_cooperation(
    db_session: AsyncSession,
):
    project = Project(name="确定性确认项目")
    db_session.add(project)
    await db_session.commit()
    await db_session.refresh(project)
    service = ExecutionService(db_session, project_id=project.id)

    saved = await service._persist_explicit_confirmation(
        [
            {
                "role": "assistant",
                "content": "计算收入时，退款订单需要扣除吗？",
                "confirmation": {
                    "key": "refund_policy",
                    "question": "计算收入时，退款订单需要扣除吗？",
                    "options": ["扣除退款", "保留退款订单"],
                },
            }
        ],
        "我确认选择“扣除退款”，请继续调查。",
    )

    assert (
        saved.items()
        >= {
            "key": "revenue_refund_policy",
            "value": "扣除退款",
            "selected_value": "扣除退款",
            "applied": True,
            "conflict": False,
            "task_query": "我确认选择“扣除退款”，请继续调查。",
        }.items()
    )
    assert saved["semantic_entry_id"]
    assert saved["active_revision_id"]
    assert len(saved["value_hash"]) == 64
    assert len(saved["definition_hash"]) == 64
    result = await db_session.execute(
        select(SemanticEntry).where(
            SemanticEntry.project_id == project.id,
            SemanticEntry.key == "revenue_refund_policy",
        )
    )
    entry = result.scalar_one()
    assert entry.state == "confirmed"
    assert entry.value == "扣除退款"
    assert entry.confidence == 1
    assert entry.source == "user"


@pytest.mark.asyncio
async def test_legacy_confirmation_reactivates_inactive_lock_and_clears_old_proof(
    db_session: AsyncSession,
):
    project = Project(name="旧锁重新确认项目")
    old = SemanticEntry(
        project=project,
        key="refund_handling",
        value="扣除退款",
        entry_type="business_rule",
        state="locked",
        confidence=1,
        definition={
            "version": 1,
            "kind": "business_rule_strategy",
            "rule_key": "refund_handling",
            "selected_option": "扣除退款",
            "action": {
                "kind": "identity",
                "column": "退款状态",
                "observed_values": ["否", "已退款"],
            },
        },
        validity="active",
        execution_state="verified",
        execution_details={"status": "verified", "proof": "old"},
        source="user",
        is_active=False,
    )
    db_session.add(project)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    receipt = await service._persist_explicit_confirmation(
        [
            {
                "role": "assistant",
                "confirmation": {
                    "key": "refund_policy",
                    "question": "计算收入时，退款订单需要扣除吗？",
                    "options": ["扣除退款", "保留退款订单"],
                },
            }
        ],
        "我确认选择“扣除退款”，请继续调查。",
    )

    await db_session.refresh(old)
    assert receipt and receipt["key"] == "revenue_refund_policy"
    assert old.key == "revenue_refund_policy"
    assert old.state == "confirmed"
    assert old.is_active is True
    assert old.definition is None
    assert old.execution_state == "definition_only"
    assert old.execution_details["status"] == "definition_only"
    assert old.revision_number == 1


@pytest.mark.asyncio
async def test_legacy_confirmation_migrates_safe_strategy_rule_key_and_revalidates(
    db_session: AsyncSession,
):
    project = Project(name="旧策略标识迁移项目")
    old = SemanticEntry(
        project=project,
        key="refund_policy",
        value="扣除退款",
        entry_type="business_rule",
        state="confirmed",
        confidence=1,
        definition={
            "version": 1,
            "kind": "business_rule_strategy",
            "rule_key": "refund_handling",
            "selected_option": "扣除退款",
            "action": {
                "kind": "identity",
                "column": "退款状态",
                "observed_values": ["否", "已退款"],
            },
        },
        validity="active",
        execution_state="verified",
        execution_details={"status": "verified", "proof": "old"},
        source="user",
    )
    db_session.add(project)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    receipt = await service._persist_explicit_confirmation(
        [
            {
                "role": "assistant",
                "confirmation": {
                    "key": "refund_policy",
                    "question": "计算收入时，退款订单需要扣除吗？",
                    "options": ["扣除退款", "保留退款订单"],
                },
            }
        ],
        "我确认选择“扣除退款”，请继续调查。",
    )

    await db_session.refresh(old)
    assert receipt and receipt["key"] == "revenue_refund_policy"
    assert old.key == "revenue_refund_policy"
    assert old.definition and old.definition["rule_key"] == "revenue_refund_policy"
    assert old.execution_state == "needs_validation"
    assert old.execution_details["status"] == "needs_validation"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "query",
    [
        "我还在扣除退款和保留退款订单之间考虑。",
        "我不选择扣除退款，请继续。",
        "继续调查，但我还没有确认答案。",
        "请分析扣除退款会怎样影响收入。",
        "我不确定是否选择扣除退款。",
    ],
)
async def test_legacy_confirmation_does_not_persist_ambiguous_or_negated_option(
    db_session: AsyncSession,
    query: str,
):
    project = Project(name="保守识别确认答案")
    db_session.add(project)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    receipt = await service._persist_explicit_confirmation(
        [
            {
                "role": "assistant",
                "confirmation": {
                    "key": "refund_policy",
                    "question": "计算收入时，退款订单需要扣除吗？",
                    "options": ["扣除退款", "保留退款订单"],
                },
            }
        ],
        query,
    )

    assert receipt is None
    entries = list(
        (
            await db_session.execute(
                select(SemanticEntry).where(SemanticEntry.project_id == project.id)
            )
        ).scalars()
    )
    assert entries == []


@pytest.mark.asyncio
async def test_legacy_confirmation_ignores_question_superseded_by_later_history(
    db_session: AsyncSession,
):
    project = Project(name="过期确认问题")
    db_session.add(project)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    receipt = await service._persist_explicit_confirmation(
        [
            {
                "role": "assistant",
                "confirmation": {
                    "key": "refund_policy",
                    "question": "计算收入时，退款订单需要扣除吗？",
                    "options": ["扣除退款", "保留退款订单"],
                },
            },
            {"role": "user", "content": "先不回答这个问题。"},
            {"role": "assistant", "content": "好的，我们先讨论其他内容。"},
        ],
        "我确认选择扣除退款。",
    )

    assert receipt is None
    entries = list(
        (
            await db_session.execute(
                select(SemanticEntry).where(SemanticEntry.project_id == project.id)
            )
        ).scalars()
    )
    assert entries == []


@pytest.mark.asyncio
async def test_project_api_upload_preflight_export_and_import(
    client: AsyncClient,
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path / "projects")
    created = await client.post(
        "/api/v1/projects",
        json={"name": "门店订单", "description": "首条黄金场景"},
    )
    assert created.status_code == 200
    project = created.json()["data"]

    csv_content = (
        "订单号,门店ID,商品品类,实付金额,退款状态\n"
        "O-1,S-1,A类,32,否\n"
        "O-1,S-1,A类,32,否\n"
        "O-2,S-2,B类,28,已退款\n"
    ).encode()
    uploaded = await client.post(
        f"/api/v1/projects/{project['id']}/sources/files",
        files={"file": ("orders.csv", csv_content, "text/csv")},
    )
    assert uploaded.status_code == 200
    source = uploaded.json()["data"]

    preflight = await client.post(
        f"/api/v1/projects/{project['id']}/sources/{source['id']}/preflight"
    )
    assert preflight.status_code == 200
    assert preflight.json()["data"]["status"] == "needs_confirmation"
    inferred_knowledge = (await client.get(f"/api/v1/projects/{project['id']}/knowledge")).json()[
        "data"
    ]
    assert not any(item["key"].startswith("grain:") for item in inferred_knowledge)
    assert not any(item["key"].startswith("metric_candidate:") for item in inferred_knowledge)
    stored_source = (await client.get(f"/api/v1/projects/{project['id']}/sources")).json()["data"][
        0
    ]
    assert stored_source["status"] == "needs_confirmation"
    first_context = await load_project_context(db_session, UUID(project["id"]))
    first_query_name = first_context.sources[0]["view_name"]

    next_month = (
        "订单号,门店ID,商品品类,实付金额,退款状态,优惠金额\n"
        "O-3,S-1,A类,30,否,2\n"
        "O-4,S-2,B类,25,否,3\n"
    ).encode()
    replacement = await client.post(
        f"/api/v1/projects/{project['id']}/sources/files",
        files={"file": ("orders-august.csv", next_month, "text/csv")},
    )
    replacement_source = replacement.json()["data"]
    replacement_preflight = await client.post(
        f"/api/v1/projects/{project['id']}/sources/{replacement_source['id']}/preflight"
    )
    replacement_report = replacement_preflight.json()["data"]
    assert any(issue["code"] == "schema_drift" for issue in replacement_report["issues"])
    current_sources = (await client.get(f"/api/v1/projects/{project['id']}/sources")).json()["data"]
    assert [item["id"] for item in current_sources] == [replacement_source["id"]]
    second_context = await load_project_context(db_session, UUID(project["id"]))
    assert len(second_context.sources) == 1
    assert second_context.sources[0]["view_name"] == first_query_name
    assert second_context.sources[0]["profile"]["version"] == 2
    learned_relationships = (
        await client.get(f"/api/v1/projects/{project['id']}/knowledge")
    ).json()["data"]
    assert not any(
        item["key"].startswith("relationship_candidate:") for item in learned_relationships
    ), "a superseded monthly version must not be learned as a separate join source"

    recipes = (await client.get(f"/api/v1/projects/{project['id']}/recipes")).json()["data"]
    replacement_recipe = next(
        recipe for recipe in recipes if recipe["data_source_id"] == replacement_source["id"]
    )
    original_recipe = next(recipe for recipe in recipes if recipe["data_source_id"] == source["id"])
    historical_reapply = await client.post(
        f"/api/v1/projects/{project['id']}/recipes/{original_recipe['id']}/reapply"
    )
    assert historical_reapply.status_code == 409
    assert replacement_recipe["operations"][0]["operation"] == "replay_prior_recipe"
    assert replacement_report["source_snapshot"]["recipe_replay"]["requested_steps"] > 0
    assert replacement_report["source_snapshot"]["recipe_replay"]["drift"] == []
    assert any(issue["code"] == "recipe_replayed" for issue in replacement_report["issues"])
    undo = await client.post(
        f"/api/v1/projects/{project['id']}/recipes/{replacement_recipe['id']}/undo"
    )
    assert undo.json()["data"]["status"] == "reverted"
    sources_after_undo = (await client.get(f"/api/v1/projects/{project['id']}/sources")).json()[
        "data"
    ]
    assert (
        next(item for item in sources_after_undo if item["id"] == replacement_source["id"])[
            "status"
        ]
        == "attached"
    )
    reapplied = await client.post(
        f"/api/v1/projects/{project['id']}/recipes/{replacement_recipe['id']}/reapply"
    )
    assert reapplied.status_code == 200
    recipes_after_reapply = (await client.get(f"/api/v1/projects/{project['id']}/recipes")).json()[
        "data"
    ]
    assert len(recipes_after_reapply) == len(recipes)
    refreshed_recipe = next(
        recipe for recipe in recipes_after_reapply if recipe["id"] == replacement_recipe["id"]
    )
    assert refreshed_recipe["status"] == "needs_attention"
    source_after_reapply = next(
        item
        for item in (await client.get(f"/api/v1/projects/{project['id']}/sources")).json()["data"]
        if item["id"] == replacement_source["id"]
    )
    assert source_after_reapply["profile_data"]["version"] == 2
    assert source_after_reapply["profile_data"]["logical_name"] == first_query_name
    assert source_after_reapply["profile_data"]["is_current"] is False
    accepted_replacement = await client.post(
        f"/api/v1/projects/{project['id']}/sources/{replacement_source['id']}/accept-replacement"
    )
    assert accepted_replacement.status_code == 200, accepted_replacement.text
    assert accepted_replacement.json()["data"]["profile_data"]["is_current"] is True

    knowledge = await client.post(
        f"/api/v1/projects/{project['id']}/knowledge",
        json={
            "key": "revenue_refund_policy",
            "value": "收入扣除退款订单",
            "state": "locked",
            "confidence": 1,
            "source": "user",
        },
    )
    assert knowledge.status_code == 200
    sources_after_confirmation = (
        await client.get(f"/api/v1/projects/{project['id']}/sources")
    ).json()["data"]
    assert all(item["status"] == "ready" for item in sources_after_confirmation)

    september = await client.post(
        f"/api/v1/projects/{project['id']}/sources/files",
        files={
            "file": (
                "orders-september.csv",
                "订单号,门店ID,实付金额,退款状态\nO-5,S-1,20,否\n".encode(),
                "text/csv",
            )
        },
    )
    september_preflight = await client.post(
        f"/api/v1/projects/{project['id']}/sources/{september.json()['data']['id']}/preflight"
    )
    september_report = september_preflight.json()["data"]
    assert september_report["status"] == "needs_confirmation"
    assert september_report["ambiguities"] == []
    assert {issue["code"] for issue in september_report["issues"]} >= {
        "recipe_replayed",
        "schema_drift",
        "replacement_pending",
    }
    assert september_report["source_snapshot"]["schema_drift"]["removed_columns"] == [
        "优惠金额",
        "商品品类",
    ]
    other_project = (await client.post("/api/v1/projects", json={"name": "另一个客户项目"})).json()[
        "data"
    ]
    other_knowledge = (
        await client.get(f"/api/v1/projects/{other_project['id']}/knowledge")
    ).json()["data"]
    assert other_knowledge == []

    analysis_run = await client.post(
        f"/api/v1/projects/{project['id']}/analysis-runs",
        json={"query": "比较各商品品类与门店的订单量"},
    )
    run = analysis_run.json()["data"]
    project_runs = (await client.get(f"/api/v1/projects/{project['id']}/analysis-runs")).json()[
        "data"
    ]
    other_runs = (await client.get(f"/api/v1/projects/{other_project['id']}/analysis-runs")).json()[
        "data"
    ]
    assert [item["id"] for item in project_runs] == [run["id"]]
    assert other_runs == []
    deleted_run = await client.delete(f"/api/v1/projects/{project['id']}/analysis-runs/{run['id']}")
    assert deleted_run.json()["data"]["deleted"] is True
    assert (await client.get(f"/api/v1/projects/{project['id']}/analysis-runs")).json()[
        "data"
    ] == []

    rejected_dependency = await client.post(
        f"/api/v1/projects/{project['id']}/dependencies",
        json={"packages": ["git+https://example.com/untrusted.git"]},
    )
    assert rejected_dependency.status_code == 422
    dependency_state = await client.get(f"/api/v1/projects/{project['id']}/dependencies")
    assert dependency_state.json()["data"] == {"requested": [], "installed": []}

    project_record = await db_session.get(Project, UUID(project["id"]))
    assert project_record is not None
    project_context = await load_project_context(db_session, project_record.id)
    golden_contract = build_golden_contract(
        query="比较各商品品类与门店的订单量",
        confirmed_knowledge=project_context.confirmed_knowledge,
        sources=project_context.sources,
        tool_history=[
            {
                "kind": "join",
                "profile": {
                    "left_key": "门店ID",
                    "right_key": "门店ID",
                    "normalization": "identifier",
                    "cardinality": "many_to_one",
                    "left_match_rate": 1,
                    "expansion_ratio": 1,
                },
            },
            {
                "kind": "validation",
                "profile": {
                    "columns": ["商品品类", "门店名称", "订单数"],
                    "keys": {"商品品类": {}, "门店名称": {}},
                    "numeric": {"订单数": {"count": 2}},
                    "truncated": False,
                },
            },
        ],
        result_rows=[{"商品品类": "A类", "门店名称": "一店", "订单数": 12}],
    )
    assert golden_contract is not None
    project_record.extra_data = {
        **(project_record.extra_data or {}),
        "golden_scenarios": [
            {
                **golden_contract,
                "id": "0123456789abcdef0123",
                "created_at": "2026-07-17T00:00:00+00:00",
                "reference_report": {
                    "metrics": [{"label": "最高组合", "value": "A类 × 一店"}],
                    "findings": ["A类与一店的关联最强"],
                },
            }
        ],
    }
    await db_session.commit()

    bundle_response = await client.get(f"/api/v1/projects/{project['id']}/export")
    bundle = bundle_response.json()["data"]
    assert any(item["state"] == "locked" for item in bundle["semantic_entries"])
    assert bundle["sanitation_histories"]
    assert bundle["golden_scenarios"][0]["query"] == "比较各商品品类与门店的订单量"

    imported = await client.post("/api/v1/projects/import", json=bundle)
    assert imported.status_code == 200
    imported_project = imported.json()["data"]
    assert imported_project["extra_data"]["recipe_template_candidates"]
    assert imported_project["extra_data"]["recipe_template_histories"]
    imported_context = await load_project_context(db_session, UUID(imported_project["id"]))
    assert imported_context.public_summary()["learned_regressions"] == 1
    restored_contract = find_matching_contract(
        imported_context.golden_scenarios,
        "比较各商品品类与门店的订单量",
    )
    assert restored_contract is not None
    assert restored_contract["result"]["required_columns"] == [
        "商品品类",
        "订单数",
        "门店名称",
    ]

    invalid_bundle = {
        **bundle,
        "project": {"name": "损坏的备份"},
        "golden_scenarios": [{**bundle["golden_scenarios"][0], "result": {}}],
    }
    rejected_import = await client.post("/api/v1/projects/import", json=invalid_bundle)
    assert rejected_import.status_code == 422


@pytest.mark.asyncio
async def test_analysis_checkpoint_round_trips_parquet_and_detects_source_drift(
    tmp_path: Path,
):
    rows = [{"门店": "一店", "订单数": 12}, {"门店": "二店", "订单数": 7}]
    sources = [
        {
            "id": "database-source",
            "kind": "connection",
            "format": "sqlite",
            "fingerprint": None,
            "profile": {
                "tables": [
                    {
                        "name": "stores",
                        "columns": [{"name": "store_id", "type": "TEXT"}],
                    }
                ]
            },
        }
    ]
    source_signatures = source_fingerprint_map(sources)
    checkpoint = await save_runtime_checkpoint(
        tmp_path,
        uuid4(),
        1,
        {
            "resumable": True,
            "source_fingerprints": source_signatures,
            "dataframes": {"store_orders": rows},
            "result_metadata": {"store_orders": {"materialized_rows": 2}},
            "validated_results": ["store_orders"],
            "replay_journal": [
                {
                    "op": "query_database",
                    "source_id": "database-source",
                    "planned_sql": "SELECT store_id FROM stores",
                    "result_name": "store_orders",
                    "result_hash": stable_payload_hash(rows),
                    "metadata_hash": stable_payload_hash({"materialized_rows": 2}),
                }
            ],
        },
    )

    restored = await load_runtime_checkpoint(tmp_path, checkpoint)

    assert restored.dataframes == {"store_orders": rows}
    assert restored.manifest["validated_results"] == ["store_orders"]
    validate_source_fingerprints(source_signatures, sources)
    changed_sources = [
        {
            **sources[0],
            "profile": {
                "tables": [
                    {
                        "name": "stores",
                        "columns": [{"name": "store_code", "type": "TEXT"}],
                    }
                ]
            },
        }
    ]
    with pytest.raises(CheckpointError):
        validate_source_fingerprints(source_signatures, changed_sources)


@pytest.mark.asyncio
async def test_checkpoint_revision_reservation_is_unique_under_parallel_tools(
    tmp_path: Path,
):
    root = tmp_path / "checkpoints"

    first, second = await asyncio.gather(
        asyncio.to_thread(_reserve_checkpoint_revision, root, 1),
        asyncio.to_thread(_reserve_checkpoint_revision, root, 1),
    )
    try:
        assert {first[0], second[0]} == {1, 2}
    finally:
        first[1].unlink(missing_ok=True)
        second[1].unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_safe_boundary_freezes_tool_state_before_checkpoint_callback(
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: TestModel())
    captured: list[dict] = []
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(name="并行检查点项目"),
    )
    runtime.deps.dataframes["first"] = [{"value": 1}]

    async def capture_checkpoint(snapshot: dict) -> dict:
        runtime.deps.dataframes["late"] = [{"value": 2}]
        captured.append(snapshot)
        return snapshot

    runtime.checkpoint_callback = capture_checkpoint
    await runtime._persist_safe_boundary()

    assert set(runtime.deps.dataframes) == {"first", "late"}
    assert set(captured[0]["dataframes"]) == {"first"}


@pytest.mark.asyncio
async def test_render_chart_checkpoint_restores_verified_artifacts_and_rejects_tampering(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    rows = [
        {"门店": "静安", "订单数": 12},
        {"门店": "西湖", "订单数": 7},
    ]
    chart_args = {
        "chart_type": "bar",
        "x": "门店",
        "y": "订单数",
        "value": None,
        "color": None,
        "title": "各门店订单量",
    }
    code = _build_result_chart_code(rows, **chart_args)
    output = "图表已生成"
    image = base64.b64encode(b"\x89PNG\r\n\x1a\nreceiptbi-chart").decode()
    checkpoint = await save_runtime_checkpoint(
        tmp_path,
        uuid4(),
        1,
        {
            "resumable": True,
            "dataframes": {"store_order_summary": rows},
            "result_metadata": {"store_order_summary": {"materialized_rows": 2}},
            "validated_results": ["store_order_summary"],
            "python_output": [output],
            "python_images": [image],
            "replay_journal": [
                {
                    "op": "query_database",
                    "purpose": "读取门店订单量数据",
                    "planned_sql": "SELECT store, orders FROM order_summary_source",
                    "result_name": "store_order_summary",
                    "result_hash": stable_payload_hash(rows),
                },
                {
                    "op": "validate_result",
                    "purpose": "核对门店订单量最终结果",
                    "result_name": "store_order_summary",
                    "result_hash": stable_payload_hash(rows),
                    "profile_hash": stable_payload_hash({"rows": 2}),
                },
                {
                    "op": "render_chart",
                    "purpose": "生成门店订单量图表",
                    "result_name": "store_order_summary",
                    **chart_args,
                    "input_hash": stable_payload_hash(rows),
                    "code_hash": stable_payload_hash(code),
                    "output_index": 0,
                    "output_hash": stable_payload_hash(output),
                    "image_start": 0,
                    "image_count": 1,
                    "image_hashes": [stable_payload_hash(image)],
                },
            ],
        },
    )

    restored = await load_runtime_checkpoint(tmp_path, checkpoint)

    assert restored.python_output == [output]
    assert restored.python_images == [image]
    assert checkpoint_manifest_is_readable(tmp_path, checkpoint) is True

    monkeypatch.setattr(analyst_runtime, "build_pydantic_model", lambda _: TestModel())
    runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(project_id=uuid4(), name="恢复图表项目"),
        resume_state={
            "manifest": restored.manifest,
            "dataframes": restored.dataframes,
            "python_output": restored.python_output,
            "python_images": restored.python_images,
        },
    )
    await runtime.replay_checkpoint()
    assert runtime.deps.python_output == [output]
    assert runtime.deps.python_images == [image]

    restored.manifest["replay_journal"][-1]["code_hash"] = "0" * 64
    drifted_runtime = PydanticAnalystRuntime(
        model_config={},
        project_context=ProjectRuntimeContext(project_id=uuid4(), name="漂移图表项目"),
        resume_state={
            "manifest": restored.manifest,
            "dataframes": restored.dataframes,
            "python_output": restored.python_output,
            "python_images": restored.python_images,
        },
    )
    with pytest.raises(CheckpointError, match="暂停时不一致"):
        await drifted_runtime.replay_checkpoint()

    manifest_path = tmp_path / str(checkpoint["manifest_path"])
    image_path = manifest_path.parent / restored.manifest["python_artifacts"]["images"][0]["path"]
    image_path.write_text(image + "A", encoding="ascii")
    with pytest.raises(CheckpointError, match="图表 1 校验失败"):
        await load_runtime_checkpoint(tmp_path, checkpoint)
    assert checkpoint_manifest_is_readable(tmp_path, checkpoint) is False


@pytest.mark.asyncio
async def test_resume_reuses_same_analysis_run_and_exposes_id_on_first_progress(
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="恢复调查项目")
    conversation = Conversation(title="门店调查", status="error")
    db_session.add_all([project, conversation])
    await db_session.flush()
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="调查各门店订单量",
        state="needs_attention",
        stage="paused",
        checkpoint={"version": 1, "revision": 1, "resumable": True},
    )
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    class Inputs:
        model_config = {"model": "test"}
        history: list[dict] = []

    class FakeEngine:
        async def execute(self, *, query, history, stop_checker):
            assert query == run.query
            yield SSEEvent.progress("understanding", "恢复工具状态")
            yield SSEEvent.result(
                "调查完成",
                report={"status": "completed", "title": "门店调查", "summary": "已完成"},
                analysis_state="completed",
                tool_history=[],
                knowledge_proposals=[],
            )

    async def fake_load_inputs(**kwargs):
        return Inputs()

    async def fake_build_engine(inputs, *, run: AnalysisRun, resume_checkpoint):
        assert run.id == run_id
        assert resume_checkpoint["resumable"] is True
        return FakeEngine()

    run_id = run.id
    monkeypatch.setattr(service, "_load_execution_inputs", fake_load_inputs)
    monkeypatch.setattr(service, "_build_engine", fake_build_engine)

    events = [
        event
        async for event in service.execute_stream(
            query="忽略这个新文案",
            conversation_id=conversation.id,
            resume_run_id=run.id,
        )
    ]

    first_progress = next(event for event in events if event.type == SSEEventType.PROGRESS)
    assert first_progress.data["analysis_run_id"] == str(run.id)
    assert first_progress.data["project_id"] == str(project.id)
    runs = list(
        (
            await db_session.execute(
                select(AnalysisRun).where(AnalysisRun.project_id == project.id)
            )
        ).scalars()
    )
    assert [item.id for item in runs] == [run.id]
    await db_session.refresh(run)
    assert run.state == "completed"
    assert run.checkpoint["resumable"] is False


@pytest.mark.asyncio
async def test_prepared_standing_run_starts_once_and_survives_restart_discovery(
    db_session: AsyncSession,
):
    project = Project(name="持续分析项目")
    conversation = Conversation(title="八月变化简报", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    standing_claim = {
        "standing_analysis_id": "standing_0123456789abcdefabcd",
        "input_token": "a" * 64,
        "idempotency_key": "b" * 64,
    }
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="重新调查各门店订单量",
        state="understanding",
        stage="prepared",
        checkpoint={"standing_analysis": standing_claim, "reason": "standing_analysis_prepared"},
    )
    db_session.add(run)
    await db_session.commit()

    assert await recover_interrupted_analysis_runs(db_session) == 0
    await db_session.refresh(run)
    assert run.state == "understanding"
    assert run.stage == "prepared"

    service = ExecutionService(db_session, project_id=project.id)
    prepared, effective_query, resume_checkpoint, receipt = await service._prepare_analysis_run(
        query="不能替换已准备的任务",
        conversation_id=conversation.id,
        resume_run_id=run.id,
    )

    assert prepared is not None and prepared.id == run.id
    assert effective_query == run.query
    assert resume_checkpoint is None
    assert receipt is None
    await db_session.refresh(run)
    assert run.stage == "understanding"
    assert run.checkpoint["standing_analysis"]["input_token"] == "a" * 64
    assert run.checkpoint["standing_analysis"]["started_at"]

    with pytest.raises(CheckpointError, match="没有可恢复"):
        await service._prepare_analysis_run(
            query=run.query,
            conversation_id=conversation.id,
            resume_run_id=run.id,
        )


@pytest.mark.asyncio
async def test_duplicate_prepared_stream_does_not_mutate_the_running_owner(
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="持续分析并发项目")
    conversation = Conversation(title="变化简报", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="重新调查收入变化",
        state="understanding",
        stage="prepared",
        checkpoint={
            "standing_analysis": {
                "id": "standing_0123456789abcdefabcd",
                "input_token": "a" * 64,
                "idempotency_key": "b" * 64,
            },
            "reason": "standing_analysis_prepared",
        },
    )
    db_session.add(run)
    await db_session.commit()
    owner = ExecutionService(db_session, project_id=project.id)
    await owner._prepare_analysis_run(
        query=run.query,
        conversation_id=conversation.id,
        resume_run_id=run.id,
    )

    class Inputs:
        model_config = {"model": "test"}
        history: list[dict] = []

    async def fake_load_inputs(**kwargs):
        return Inputs()

    duplicate = ExecutionService(db_session, project_id=project.id)
    monkeypatch.setattr(duplicate, "_load_execution_inputs", fake_load_inputs)
    events = [
        event
        async for event in duplicate.execute_stream(
            query=run.query,
            conversation_id=conversation.id,
            resume_run_id=run.id,
        )
    ]

    assert [event.type for event in events] == [SSEEventType.ERROR]
    assert events[0].data["code"] == "RESUME_UNAVAILABLE"
    await db_session.refresh(run)
    assert run.state == "understanding"
    assert run.stage == "understanding"
    assert run.checkpoint["standing_analysis"]["input_token"] == "a" * 64


@pytest.mark.asyncio
async def test_missing_data_resume_reuses_same_run_without_replaying_old_tools(
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="补数继续调查项目")
    conversation = Conversation(title="门店订单调查", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="比较各商品品类与门店的订单量",
        state="needs_attention",
        stage="needs_attention",
        checkpoint={
            "resumable": True,
            "reason": "awaiting_data",
            "tool_history": [{"kind": "inspect_project"}],
        },
        report={
            "status": "needs_data",
            "title": "还需要门店资料",
            "action": {
                "kind": "add_data",
                "label": "补充门店资料",
                "requested_data": ["门店编号和门店名称"],
            },
        },
    )
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    class Inputs:
        model_config = {"model": "test"}
        history: list[dict] = []

    class FakeEngine:
        async def execute(self, *, query, history, stop_checker):
            assert query == run.query
            yield SSEEvent.result(
                "调查完成",
                report={"status": "completed", "title": "门店订单量比较", "summary": "已完成"},
                analysis_state="completed",
                tool_history=[],
                knowledge_proposals=[],
            )

    async def fake_load_inputs(**kwargs):
        return Inputs()

    async def fake_build_engine(inputs, *, run: AnalysisRun, resume_checkpoint):
        assert run.id == run_id
        assert resume_checkpoint is None
        return FakeEngine()

    run_id = run.id
    monkeypatch.setattr(service, "_load_execution_inputs", fake_load_inputs)
    monkeypatch.setattr(service, "_build_engine", fake_build_engine)

    events = [
        event
        async for event in service.execute_stream(
            query="数据已补好，继续调查",
            conversation_id=conversation.id,
            resume_run_id=run.id,
        )
    ]

    result = next(event for event in events if event.type == SSEEventType.RESULT)
    assert result.data["analysis_run_id"] == str(run.id)
    runs = list(
        (
            await db_session.execute(
                select(AnalysisRun).where(AnalysisRun.project_id == project.id)
            )
        ).scalars()
    )
    assert [item.id for item in runs] == [run.id]
    await db_session.refresh(run)
    assert run.state == "completed"
    assert run.checkpoint["reason"] == "completed"
    assert run.checkpoint["resumable"] is False


@pytest.mark.asyncio
async def test_missing_data_continuation_stays_retryable_until_a_result_is_saved(
    db_session: AsyncSession,
):
    project = Project(name="补数恢复保护项目")
    conversation = Conversation(title="补数调查", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="调查门店销售",
        state="needs_attention",
        stage="needs_attention",
        checkpoint={"resumable": True, "reason": "awaiting_data"},
        report={"status": "needs_data", "title": "需要门店资料"},
    )
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    prepared, query, checkpoint, receipt = await service._prepare_analysis_run(
        query="数据已补好",
        conversation_id=conversation.id,
        resume_run_id=run.id,
    )
    assert prepared is run
    assert query == run.query
    assert checkpoint is None
    assert receipt is None
    assert run.checkpoint["continuation_kind"] == "data"
    assert run.checkpoint["resumable"] is True

    await service._mark_run_needs_attention(run, RuntimeError("模型连接暂时失败"))
    await db_session.refresh(run)
    assert run.state == "needs_attention"
    assert run.checkpoint["continuation_kind"] == "data"
    assert run.checkpoint["resumable"] is True

    retried, retried_query, retried_checkpoint, _ = await service._prepare_analysis_run(
        query="继续",
        conversation_id=conversation.id,
        resume_run_id=run.id,
    )
    assert retried is run
    assert retried_query == run.query
    assert retried_checkpoint is None


@pytest.mark.asyncio
async def test_startup_recovery_marks_interrupted_run_and_writes_resume_message(
    db_session: AsyncSession,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    monkeypatch.setattr(settings, "WORKSPACE_ROOT", tmp_path)
    project = Project(name="启动恢复项目")
    conversation = Conversation(title="未完成调查", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    rows = [{"门店": "一店", "订单数": 12}]
    checkpoint = await save_runtime_checkpoint(
        tmp_path / str(project.id),
        uuid4(),
        1,
        {
            "resumable": True,
            "source_fingerprints": {},
            "dataframes": {"store_orders": rows},
            "replay_journal": [
                {
                    "op": "query_project_files",
                    "planned_sql": "SELECT * FROM orders",
                    "result_name": "store_orders",
                    "result_hash": stable_payload_hash(rows),
                    "metadata_hash": stable_payload_hash({}),
                }
            ],
        },
    )
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="调查门店",
        state="investigating",
        stage="investigating",
        checkpoint=checkpoint,
    )
    db_session.add(run)
    await db_session.commit()

    recovered = await recover_interrupted_analysis_runs(db_session)
    await db_session.commit()

    assert recovered == 1
    await db_session.refresh(run)
    assert run.state == "needs_attention"
    assert run.checkpoint["resumable"] is True
    message_result = await db_session.execute(
        select(Message).where(Message.conversation_id == conversation.id)
    )
    message = message_result.scalar_one()
    assert message.extra_data["analysis_run_id"] == str(run.id)
    assert message.extra_data["resumable"] is True
    assert message.extra_data["error_code"] == "PROCESS_INTERRUPTED"


@pytest.mark.asyncio
async def test_terminal_events_are_not_emitted_when_result_persistence_fails(
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="结果耐久性项目")
    conversation = Conversation(title="持久化失败调查", status="active")
    db_session.add_all([project, conversation])
    await db_session.flush()
    run = AnalysisRun(
        project_id=project.id,
        conversation_id=conversation.id,
        query="调查收入变化",
        state="needs_attention",
        stage="paused",
        checkpoint={"version": 1, "revision": 1, "resumable": True},
    )
    db_session.add(run)
    await db_session.commit()
    service = ExecutionService(db_session, project_id=project.id)

    class Inputs:
        model_config = {"model": "test"}
        history: list[dict] = []

    class FakeEngine:
        async def execute(self, *, query, history, stop_checker):
            yield SSEEvent.progress("completed", "调查完成，正在保存")
            yield SSEEvent.result(
                "调查完成",
                report={"status": "completed", "title": "收入变化", "summary": "已完成"},
                analysis_state="completed",
                tool_history=[],
                knowledge_proposals=[],
            )

    async def fake_load_inputs(**kwargs):
        return Inputs()

    async def fake_build_engine(inputs, *, run: AnalysisRun, resume_checkpoint):
        return FakeEngine()

    async def fail_after_staging(run: AnalysisRun, result_data: dict) -> None:
        run.state = "completed"
        db_session.add(
            ArtifactRecord(
                project_id=run.project_id,
                analysis_run_id=run.id,
                kind="report",
                title="不应提交的报告",
                payload={"status": "completed"},
                technical_details={},
            )
        )
        raise RuntimeError("forced persistence failure")

    monkeypatch.setattr(service, "_load_execution_inputs", fake_load_inputs)
    monkeypatch.setattr(service, "_build_engine", fake_build_engine)
    monkeypatch.setattr(service, "_persist_project_result", fail_after_staging)

    events = [
        event
        async for event in service.execute_stream(
            query="继续",
            conversation_id=conversation.id,
            resume_run_id=run.id,
        )
    ]

    assert [event.type for event in events] == [SSEEventType.ERROR]
    assert events[0].data["code"] == "RUNTIME_ERROR"
    await db_session.refresh(run)
    assert run.state == "needs_attention"
    artifacts = list(
        (
            await db_session.execute(
                select(ArtifactRecord).where(ArtifactRecord.analysis_run_id == run.id)
            )
        ).scalars()
    )
    assert artifacts == []
