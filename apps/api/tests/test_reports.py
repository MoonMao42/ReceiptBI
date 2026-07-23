"""Editable report document API tests."""

from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import quote
from uuid import uuid4

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete, func, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.v1 import reports as reports_api
from app.db.tables import (
    AnalysisRun,
    ArtifactRecord,
    Project,
    ReportBlock,
    ReportDocument,
    ReportPage,
)
from app.models.workspace import AnalysisPlaybookResponse
from app.services import analysis_playbook_runner
from app.services.analysis_playbook_runner import (
    AnalysisPlaybookExecutionReceipt,
    AnalysisPlaybookRunnerError,
    AnalysisPlaybookRunResult,
)


async def _seed_analysis(
    db_session: AsyncSession,
    project: Project,
) -> tuple[AnalysisRun, ArtifactRecord]:
    run = AnalysisRun(
        project_id=project.id,
        query="概览本月经营表现",
        state="completed",
        stage="completed",
        report={"summary": "本月收入增长"},
    )
    db_session.add(run)
    await db_session.flush()
    artifact = ArtifactRecord(
        project_id=project.id,
        analysis_run_id=run.id,
        kind="chart",
        title="月度收入趋势",
        payload={"series": [100, 120, 140]},
        technical_details={"query_hash": "immutable"},
    )
    db_session.add(artifact)
    await db_session.commit()
    return run, artifact


def _refresh_playbook(
    *,
    playbook_id: str = "pb_0123456789abcdefabcd",
    query: str = "按区域汇总收入",
    limit: int = 100,
) -> AnalysisPlaybookResponse:
    now = datetime.now(UTC)
    payload = {
        "schema_version": 3,
        "execution_mode": "system_structured_query",
        "id": playbook_id,
        "name": "区域收入",
        "query": query,
        "source_roles": [
            {
                "logical_name": "orders",
                "source_kind": "file",
                "tables": ["orders"],
                "columns": [
                    {
                        "name": "region",
                        "data_type": "VARCHAR",
                        "canonical_type": "text",
                    },
                    {
                        "name": "revenue",
                        "data_type": "DOUBLE",
                        "canonical_type": "number",
                    },
                ],
                "schema_signature": "a" * 64,
            }
        ],
        "confirmed_knowledge_keys": [],
        "relationship_keys": [],
        "steps": [
            {
                "order": 1,
                "kind": "structured_query",
                "summary": "按区域汇总收入",
                "input_results": [],
                "output_result": "result_1",
                "source_role": "orders",
                "plan": {
                    "table": "orders",
                    "dimensions": ["region"],
                    "metrics": [
                        {
                            "operation": "sum",
                            "column": "revenue",
                            "alias": "revenue",
                        }
                    ],
                    "filters": [],
                    "sort": [{"field": "revenue", "direction": "desc"}],
                    "limit": limit,
                },
            },
            {
                "order": 2,
                "kind": "validate_result",
                "summary": "校验区域收入",
                "input_results": ["result_1"],
                "output_result": None,
                "key_columns": ["region"],
                "numeric_columns": ["revenue"],
                "must_not_be_truncated": True,
            },
        ],
        "validation": {
            "input_result": "result_1",
            "columns": ["region", "revenue"],
            "key_columns": ["region"],
            "numeric_columns": ["revenue"],
            "must_not_be_truncated": True,
        },
        "shape_hash": "0" * 64,
        "created_at": now,
        "updated_at": now,
    }
    playbook = AnalysisPlaybookResponse.model_validate(payload)
    return playbook.model_copy(
        update={"shape_hash": analysis_playbook_runner._shape_hash(playbook)}
    )


async def _seed_refreshable_report(
    db: AsyncSession,
    *,
    block_type: str = "table",
    artifact_payload: dict | None = None,
    content: dict | None = None,
    config: dict | None = None,
    layout: dict | None = None,
) -> tuple[Project, AnalysisPlaybookResponse, AnalysisRun, ArtifactRecord, ReportDocument]:
    query = "按区域汇总收入"
    playbook = _refresh_playbook(query=query)
    project = Project(
        name="可刷新报表",
        extra_data={"analysis_playbooks": [playbook.model_dump(mode="json")]},
    )
    db.add(project)
    await db.flush()
    source_ref = {
        "source_id": str(uuid4()),
        "source_logical_name": "orders",
        "source_kind": "file",
        "table_or_view": "orders",
        "query_scope": "aggregated",
    }
    query_plan = playbook.steps[0].plan.model_dump(mode="json")
    tool_history = [
        {
            "kind": "structured_query",
            "source_kind": "file",
            "source_id": source_ref["source_id"],
            "source_refs": [source_ref],
            "purpose": "按区域汇总收入",
            "query_plan": query_plan,
            "compiled_sql": "SELECT region, SUM(revenue) AS revenue FROM orders GROUP BY region",
            "result_name": "result_1",
            "rows": 1,
            "truncated": False,
            "result_completeness": "complete",
        },
        {
            "kind": "validation",
            "purpose": "校验区域收入",
            "result_name": "result_1",
            "profile": {
                "materialized_rows": 1,
                "columns": ["region", "revenue"],
                "keys": {"region": {"unique": 1}},
                "numeric": {"revenue": {"count": 1}},
                "truncated": False,
                "source_refs": [source_ref],
            },
        },
    ]
    run = AnalysisRun(
        project_id=project.id,
        query=query,
        state="completed",
        stage="completed",
        report={"title": "区域收入"},
        checkpoint={"tool_history": tool_history},
    )
    db.add(run)
    await db.flush()
    if artifact_payload is None:
        artifact_payload = (
            {
                "chart": {
                    "version": 1,
                    "type": "bar",
                    "title": "区域收入",
                    "data_ref": {"result_name": "result_1", "result_hash": "b" * 64},
                    "encoding": {
                        "x": {"field": "region"},
                        "y": [{"field": "revenue", "format": "currency"}],
                    },
                    "presentation": {
                        "orientation": "vertical",
                        "stack": "none",
                        "palette": "receiptbi",
                    },
                    "data": [{"region": "旧区域", "revenue": 1}],
                }
            }
            if block_type == "chart"
            else {
                "rows": [{"region": "旧区域", "revenue": 1}],
                "rows_count": 1,
                "sampled": False,
                "columns": ["region", "revenue"],
            }
        )
    artifact = ArtifactRecord(
        project_id=project.id,
        analysis_run_id=run.id,
        kind="chart" if block_type == "chart" else "table",
        title="区域收入",
        payload=artifact_payload,
        technical_details={"result_name": "result_1"},
    )
    db.add(artifact)
    await db.flush()
    page = ReportPage(title="概览", order_index=0, config={}, blocks=[])
    report = ReportDocument(
        project_id=project.id,
        title="经营报告",
        status="draft",
        extra_data={},
        pages=[page],
    )
    page.blocks.append(
        ReportBlock(
            block_type=block_type,
            title="人工保留标题",
            order_index=0,
            source_kind="artifact",
            analysis_run_id=run.id,
            artifact_id=artifact.id,
            source_ref={"snapshot": {"title": "历史快照"}},
            content=content if content is not None else dict(artifact_payload),
            config=config or {"manual_override": True},
            layout=layout or {"x": 2, "y": 3, "w": 8, "h": 5},
        )
    )
    db.add(report)
    await db.commit()
    return project, playbook, run, artifact, report


def _fake_refresh_result(
    playbook: AnalysisPlaybookResponse,
    rows: list[dict],
    *,
    receipt_playbook_id: str | None = None,
) -> AnalysisPlaybookRunResult:
    receipt = AnalysisPlaybookExecutionReceipt(
        playbook_id=receipt_playbook_id or playbook.id,
        playbook_shape_hash=playbook.shape_hash,
        source_role="orders",
        source_kind="file",
        source_id="source-1",
        source_schema_signature="a" * 64,
        plan_hash="b" * 64,
        result_name="result_1",
        row_count=len(rows),
        execution_backend="duckdb",
        result_hash="c" * 64,
        metadata_hash="d" * 64,
        profile_hash="e" * 64,
        validation_hash="f" * 64,
    )
    return AnalysisPlaybookRunResult(
        result_name="result_1",
        rows=rows,
        metadata={},
        tool_history=[],
        replay_journal=[],
        validated_results={"result_1"},
        validation={},
        receipt=receipt,
    )


def _stub_refresh_runtime(
    monkeypatch: pytest.MonkeyPatch,
    result: AnalysisPlaybookRunResult,
) -> None:
    async def load_context(*_args, **_kwargs):
        return SimpleNamespace(
            sources=[],
            project_dir=Path("/tmp"),
            connection_configs={},
        )

    async def run_playbook(*_args, **_kwargs):
        return result

    monkeypatch.setattr(reports_api, "load_project_context", load_context)
    monkeypatch.setattr(reports_api, "run_analysis_playbook", run_playbook)
    monkeypatch.setattr(
        reports_api,
        "validate_playbook_execution_evidence",
        lambda *_args, **_kwargs: None,
    )


@pytest.mark.asyncio
async def test_create_list_and_get_report_with_manual_and_artifact_blocks(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="经营驾驶舱")
    db_session.add(project)
    await db_session.flush()
    run, artifact = await _seed_analysis(db_session, project)

    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "七月经营报告",
            "description": "人工叙事与已验证调查结果的组合",
            "pages": [
                {
                    "title": "概览",
                    "order_index": 0,
                    "config": {"columns": 12},
                    "blocks": [
                        {
                            "block_type": "text",
                            "title": "管理摘要",
                            "order_index": 0,
                            "source_kind": "manual",
                            "content": {"text": "这里可以自由编辑。"},
                            "layout": {"x": 0, "y": 0, "w": 12, "h": 3},
                        },
                        {
                            "block_type": "chart",
                            "title": "收入趋势",
                            "order_index": 1,
                            "source_kind": "artifact",
                            "artifact_id": str(artifact.id),
                            "content": {"caption": "来自已完成调查"},
                            "layout": {"x": 0, "y": 3, "w": 8, "h": 6},
                            "config": {"chart_type": "line"},
                        },
                    ],
                }
            ],
        },
    )

    assert created.status_code == 201, created.text
    document = created.json()["data"]
    assert document["version"] == 1
    assert document["pages"][0]["config"] == {"columns": 12}
    assert [block["source_kind"] for block in document["pages"][0]["blocks"]] == [
        "manual",
        "artifact",
    ]
    artifact_block = document["pages"][0]["blocks"][1]
    assert artifact_block["artifact_id"] == str(artifact.id)
    assert artifact_block["analysis_run_id"] == str(run.id)
    assert artifact_block["source_available"] is True
    assert artifact_block["source_ref"]["artifact_id"] == str(artifact.id)
    assert artifact_block["source_ref"]["snapshot"]["payload"] == {
        "series": [100, 120, 140]
    }

    listing = await client.get(f"/api/v1/projects/{project.id}/reports")
    assert listing.status_code == 200
    assert listing.json()["data"] == [
        {
            "id": document["id"],
            "project_id": str(project.id),
            "title": "七月经营报告",
            "description": "人工叙事与已验证调查结果的组合",
            "status": "draft",
            "version": 1,
            "page_count": 1,
            "block_count": 2,
            "created_at": document["created_at"],
            "updated_at": document["updated_at"],
        }
    ]

    detail = await client.get(
        f"/api/v1/projects/{project.id}/reports/{document['id']}"
    )
    assert detail.status_code == 200
    assert detail.json()["data"] == document
    stored_artifact = await db_session.get(ArtifactRecord, artifact.id)
    assert stored_artifact is not None
    assert stored_artifact.payload == {"series": [100, 120, 140]}
    assert stored_artifact.technical_details == {"query_hash": "immutable"}


@pytest.mark.asyncio
async def test_full_document_update_upserts_pages_and_blocks_with_version_guard(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="可编辑报表")
    db_session.add(project)
    await db_session.commit()
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "初稿",
            "pages": [
                {
                    "title": "概览",
                    "blocks": [
                        {
                            "block_type": "text",
                            "source_kind": "manual",
                            "content": {"text": "旧内容"},
                            "layout": {"x": 0, "y": 0, "w": 6, "h": 2},
                        }
                    ],
                }
            ],
        },
    )
    initial = created.json()["data"]
    page = initial["pages"][0]
    block = page["blocks"][0]

    updated = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{initial['id']}",
        json={
            "expected_version": 1,
            "title": "正式经营报告",
            "pages": [
                {
                    "id": page["id"],
                    "title": "经营概览",
                    "order_index": 0,
                    "config": {"background": "paper"},
                    "blocks": [
                        {
                            "id": block["id"],
                            "block_type": "text",
                            "title": "结论",
                            "order_index": 1,
                            "source_kind": "manual",
                            "content": {"text": "用户修改后的内容"},
                            "layout": {"x": 0, "y": 0, "w": 12, "h": 2},
                            "config": {"tone": "plain"},
                        },
                        {
                            "block_type": "metric",
                            "title": "手工目标",
                            "order_index": 0,
                            "source_kind": "manual",
                            "content": {"value": 180, "unit": "万元"},
                            "layout": {"x": 0, "y": 2, "w": 3, "h": 2},
                        },
                    ],
                },
                {
                    "title": "明细",
                    "order_index": 1,
                    "blocks": [],
                },
            ],
        },
    )

    assert updated.status_code == 200, updated.text
    document = updated.json()["data"]
    assert document["title"] == "正式经营报告"
    assert document["version"] == 2
    assert [page["title"] for page in document["pages"]] == ["经营概览", "明细"]
    assert document["pages"][0]["version"] == 2
    assert document["pages"][0]["blocks"][1]["id"] == block["id"]
    assert document["pages"][0]["blocks"][1]["version"] == 2
    assert document["pages"][0]["blocks"][1]["content"] == {
        "text": "用户修改后的内容"
    }

    # Bypass the in-memory fast-fail: the database CAS must still reject the
    # second writer that presents the same expected version.
    monkeypatch.setattr(reports_api, "_assert_expected_version", lambda *_args: None)
    stale = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{initial['id']}",
        json={"expected_version": 1, "title": "过期覆盖"},
    )
    assert stale.status_code == 409
    detail = await client.get(
        f"/api/v1/projects/{project.id}/reports/{initial['id']}"
    )
    assert detail.json()["data"]["title"] == "正式经营报告"


@pytest.mark.asyncio
async def test_report_source_references_are_project_scoped(
    client: AsyncClient,
    db_session: AsyncSession,
):
    report_project = Project(name="报表项目")
    foreign_project = Project(name="其他项目")
    db_session.add_all([report_project, foreign_project])
    await db_session.flush()
    _, foreign_artifact = await _seed_analysis(db_session, foreign_project)

    rejected = await client.post(
        f"/api/v1/projects/{report_project.id}/reports",
        json={
            "title": "不能跨项目引用",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "chart",
                            "source_kind": "artifact",
                            "artifact_id": str(foreign_artifact.id),
                        }
                    ]
                }
            ],
        },
    )

    assert rejected.status_code == 400
    assert "不属于当前项目" in rejected.json()["detail"]
    report_count = await db_session.scalar(select(func.count(ReportDocument.id)))
    assert report_count == 0


@pytest.mark.asyncio
async def test_delete_report_cascades_editable_tree_but_keeps_analysis_artifact(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="报表删除边界")
    db_session.add(project)
    await db_session.flush()
    run, artifact = await _seed_analysis(db_session, project)
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "待删除报表",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "chart",
                            "source_kind": "artifact",
                            "artifact_id": str(artifact.id),
                        }
                    ]
                }
            ],
        },
    )
    report_id = created.json()["data"]["id"]

    deleted = await client.delete(
        f"/api/v1/projects/{project.id}/reports/{report_id}"
    )

    assert deleted.status_code == 200
    assert deleted.json()["data"] == {"id": report_id, "deleted": True}
    assert await db_session.scalar(select(func.count(ReportDocument.id))) == 0
    assert await db_session.scalar(select(func.count(ReportPage.id))) == 0
    assert await db_session.scalar(select(func.count(ReportBlock.id))) == 0
    assert await db_session.get(AnalysisRun, run.id) is not None
    assert await db_session.get(ArtifactRecord, artifact.id) is not None


@pytest.mark.asyncio
async def test_block_crud_keeps_manual_source_explicit(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="区块编辑")
    db_session.add(project)
    await db_session.commit()
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={"title": "区块编辑测试"},
    )
    report = created.json()["data"]
    page = report["pages"][0]

    added = await client.post(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks",
        json={
            "block_type": "text",
            "source_kind": "manual",
            "content": {"text": "用户输入"},
        },
    )
    assert added.status_code == 201
    block = added.json()["data"]["pages"][0]["blocks"][0]
    assert block["source_kind"] == "manual"
    assert block["version"] == 1

    edited = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks/{block['id']}",
        json={
            "expected_version": 1,
            "title": "手工备注",
            "content": {"text": "可以继续编辑"},
            "layout": {"x": 3, "y": 1, "w": 6, "h": 2},
        },
    )
    assert edited.status_code == 200
    edited_block = edited.json()["data"]["pages"][0]["blocks"][0]
    assert edited_block["title"] == "手工备注"
    assert edited_block["version"] == 2
    assert edited_block["source_kind"] == "manual"

    removed = await client.delete(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks/{block['id']}"
    )
    assert removed.status_code == 200
    assert removed.json()["data"]["pages"][0]["blocks"] == []


@pytest.mark.asyncio
async def test_patch_rejects_null_for_non_nullable_report_fields(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="拒绝空值")
    db_session.add(project)
    await db_session.commit()
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "保持有效",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "text",
                            "source_kind": "manual",
                            "content": {"text": "不会被空值覆盖"},
                        }
                    ]
                }
            ],
        },
    )
    report = created.json()["data"]
    page = report["pages"][0]
    block = page["blocks"][0]

    invalid_report = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}",
        json={"expected_version": report["version"], "title": None},
    )
    invalid_page = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}",
        json={"expected_version": page["version"], "config": None},
    )
    invalid_block = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks/{block['id']}",
        json={"expected_version": block["version"], "content": None},
    )

    assert invalid_report.status_code == 422
    assert invalid_page.status_code == 422
    assert invalid_block.status_code == 422
    detail = await client.get(
        f"/api/v1/projects/{project.id}/reports/{report['id']}"
    )
    assert detail.json()["data"]["title"] == "保持有效"
    assert detail.json()["data"]["pages"][0]["config"] == {}
    assert detail.json()["data"]["pages"][0]["blocks"][0]["content"] == {
        "text": "不会被空值覆盖"
    }


@pytest.mark.asyncio
async def test_page_and_block_patch_use_database_compare_and_swap(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project = Project(name="并发编辑")
    db_session.add(project)
    await db_session.commit()
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "并发保护",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "text",
                            "source_kind": "manual",
                            "content": {"text": "原始内容"},
                        }
                    ]
                }
            ],
        },
    )
    report = created.json()["data"]
    page = report["pages"][0]
    block = page["blocks"][0]

    first_page_write = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}",
        json={"expected_version": page["version"], "title": "第一个写入者"},
    )
    assert first_page_write.status_code == 200

    # Ignore the in-memory precheck; UPDATE ... WHERE id/version must be the
    # layer that rejects a second writer with the same version.
    monkeypatch.setattr(reports_api, "_assert_expected_version", lambda *_args: None)
    stale_page_write = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}",
        json={"expected_version": page["version"], "title": "第二个写入者"},
    )
    assert stale_page_write.status_code == 409

    first_block_write = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks/{block['id']}",
        json={
            "expected_version": block["version"],
            "content": {"text": "第一个区块写入者"},
        },
    )
    assert first_block_write.status_code == 200
    stale_block_write = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report['id']}/pages/{page['id']}/blocks/{block['id']}",
        json={
            "expected_version": block["version"],
            "content": {"text": "第二个区块写入者"},
        },
    )
    assert stale_block_write.status_code == 409

    detail = await client.get(
        f"/api/v1/projects/{project.id}/reports/{report['id']}"
    )
    current = detail.json()["data"]
    assert current["pages"][0]["title"] == "第一个写入者"
    assert current["pages"][0]["blocks"][0]["content"] == {
        "text": "第一个区块写入者"
    }


@pytest.mark.asyncio
async def test_source_snapshot_survives_analysis_cleanup_and_remains_editable(
    client: AsyncClient,
    db_session: AsyncSession,
):
    await db_session.execute(text("PRAGMA foreign_keys=ON"))
    await db_session.commit()
    project = Project(name="来源保留")
    db_session.add(project)
    await db_session.flush()
    run, artifact = await _seed_analysis(db_session, project)
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "独立于调查历史的报表",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "chart",
                            "source_kind": "artifact",
                            "artifact_id": str(artifact.id),
                            "content": {"caption": "已经固定到报表的静态说明"},
                        }
                    ]
                }
            ],
        },
    )
    document = created.json()["data"]
    original_block = document["pages"][0]["blocks"][0]
    original_ref = original_block["source_ref"]

    # This is the same cleanup boundary used when an investigation is removed:
    # artifacts cascade away and live report FKs become NULL, while the report
    # itself must remain readable.
    await db_session.execute(delete(AnalysisRun).where(AnalysisRun.id == run.id))
    await db_session.commit()

    detail = await client.get(
        f"/api/v1/projects/{project.id}/reports/{document['id']}"
    )
    detached = detail.json()["data"]
    detached_page = detached["pages"][0]
    detached_block = detached_page["blocks"][0]
    assert detached_block["analysis_run_id"] is None
    assert detached_block["artifact_id"] is None
    assert detached_block["source_available"] is False
    assert detached_block["source_ref"] == original_ref
    assert detached_block["source_ref"]["artifact_id"] == str(artifact.id)
    assert detached_block["source_ref"]["snapshot"]["payload"] == {
        "series": [100, 120, 140]
    }
    assert detached_block["content"] == {
        "caption": "已经固定到报表的静态说明"
    }

    edited = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{document['id']}",
        json={
            "expected_version": detached["version"],
            "pages": [
                {
                    "id": detached_page["id"],
                    "title": detached_page["title"],
                    "order_index": detached_page["order_index"],
                    "config": detached_page["config"],
                    "blocks": [
                        {
                            "id": detached_block["id"],
                            "block_type": detached_block["block_type"],
                            "title": detached_block["title"],
                            "order_index": detached_block["order_index"],
                            "source_kind": detached_block["source_kind"],
                            "analysis_run_id": None,
                            "artifact_id": None,
                            "source_ref": detached_block["source_ref"],
                            "content": {"caption": "来源已清理，但内容仍可编辑"},
                            "layout": detached_block["layout"],
                            "config": detached_block["config"],
                        }
                    ],
                }
            ],
        },
    )
    assert edited.status_code == 200, edited.text
    edited_block = edited.json()["data"]["pages"][0]["blocks"][0]
    assert edited_block["source_available"] is False
    assert edited_block["source_ref"] == original_ref
    assert edited_block["content"] == {"caption": "来源已清理，但内容仍可编辑"}


@pytest.mark.asyncio
async def test_unknown_block_type_is_rejected_before_persistence(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="区块类型契约")
    db_session.add(project)
    await db_session.commit()

    rejected = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": "不可渲染区块",
            "pages": [
                {
                    "blocks": [
                        {
                            "block_type": "narrative",
                            "source_kind": "manual",
                        }
                    ]
                }
            ],
        },
    )

    assert rejected.status_code == 422
    assert await db_session.scalar(select(func.count(ReportDocument.id))) == 0


def test_report_source_kind_requires_explicit_provenance() -> None:
    # FastAPI/Pydantic handles malformed provenance before touching the database.
    from app.models.reports import ReportBlockCreate

    with pytest.raises(ValueError):
        ReportBlockCreate(block_type="text", source_kind="manual", analysis_run_id=uuid4())


@pytest.mark.asyncio
async def test_refresh_table_rebinds_current_rows_and_preserves_manual_configuration(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    old_content = {
        "rows": [{"region": "旧区域", "revenue": 1}],
        "rows_count": 1,
        "sampled": False,
        "columns": ["region", "revenue"],
        "note": "保留的人工说明",
        "_filter_applied": True,
        "_filter_source_rows": 9,
    }
    project, playbook, _, _, report = await _seed_refreshable_report(
        db_session,
        content=old_content,
        config={"manual_override": True, "number_format": "currency"},
    )
    page = report.pages[0]
    block = page.blocks[0]
    rows = [{"region": f"区域-{index}", "revenue": index} for index in range(205)]
    _stub_refresh_runtime(monkeypatch, _fake_refresh_result(playbook, rows))

    refreshed = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert refreshed.status_code == 200, refreshed.text
    document = refreshed.json()["data"]
    current_page = document["pages"][0]
    current = current_page["blocks"][0]
    assert document["version"] == 2
    assert current_page["version"] == 2
    assert current["version"] == 2
    assert current["title"] == "人工保留标题"
    assert current["layout"] == {"x": 2, "y": 3, "w": 8, "h": 5}
    assert current["config"] == {
        "manual_override": True,
        "number_format": "currency",
    }
    assert current["content"]["rows"] == rows[:200]
    assert current["content"]["rows_count"] == 205
    assert current["content"]["sampled"] is True
    assert current["content"]["columns"] == ["region", "revenue"]
    assert current["content"]["note"] == "保留的人工说明"
    assert "_filter_applied" not in current["content"]
    assert "_filter_source_rows" not in current["content"]
    assert current["source_ref"]["snapshot"] == {"title": "历史快照"}
    assert current["source_ref"]["refresh_binding"] == {
        "version": 1,
        "kind": "analysis_playbook",
        "playbook_id": playbook.id,
        "playbook_shape_hash": playbook.shape_hash,
        "result_name": "result_1",
    }


@pytest.mark.asyncio
async def test_refresh_chart_only_rebinds_chart_data_and_data_reference(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    chart = {
        "version": 1,
        "type": "bar",
        "title": "用户修改后的图表标题",
        "data_ref": {"result_name": "result_1", "result_hash": "1" * 64},
        "encoding": {
            "x": {"field": "region"},
            "y": [{"field": "revenue", "format": "currency"}],
        },
        "presentation": {
            "orientation": "horizontal",
            "stack": "none",
            "palette": "monochrome",
        },
        "data": [{"region": "旧区域", "revenue": 1}],
    }
    project, playbook, _, _, report = await _seed_refreshable_report(
        db_session,
        block_type="chart",
        content={
            "chart": chart,
            "rows": chart["data"],
            "caption": "保留说明",
            "_filter_applied": True,
        },
        config={"manual_override": True, "show_labels": True},
    )
    page = report.pages[0]
    block = page.blocks[0]
    rows = [
        {"region": "华东", "revenue": 120},
        {"region": "华南", "revenue": 80},
    ]
    _stub_refresh_runtime(monkeypatch, _fake_refresh_result(playbook, rows))

    refreshed = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert refreshed.status_code == 200, refreshed.text
    current = refreshed.json()["data"]["pages"][0]["blocks"][0]
    assert current["config"] == {"manual_override": True, "show_labels": True}
    assert current["content"]["caption"] == "保留说明"
    assert current["content"]["rows"] == rows
    assert "_filter_applied" not in current["content"]
    rebound_chart = current["content"]["chart"]
    assert rebound_chart["title"] == "用户修改后的图表标题"
    assert rebound_chart["encoding"] == chart["encoding"]
    assert rebound_chart["presentation"] == chart["presentation"]
    assert rebound_chart["data"] == rows
    assert rebound_chart["data_ref"]["result_name"] == "result_1"
    assert rebound_chart["data_ref"]["result_hash"] != chart["data_ref"]["result_hash"]


@pytest.mark.asyncio
async def test_refresh_uses_database_compare_and_swap_after_execution(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    page = report.pages[0]
    block = page.blocks[0]
    project_id = project.id
    report_id = report.id
    refresh_url = (
        f"/api/v1/projects/{project_id}/reports/{report_id}/pages/{page.id}"
        f"/blocks/{block.id}/refresh"
    )
    _stub_refresh_runtime(
        monkeypatch,
        _fake_refresh_result(playbook, [{"region": "华东", "revenue": 120}]),
    )
    first = await client.post(
        refresh_url,
        json={"expected_version": 1},
    )
    assert first.status_code == 200

    monkeypatch.setattr(reports_api, "_assert_expected_version", lambda *_args: None)
    stale = await client.post(
        refresh_url,
        json={"expected_version": 1},
    )

    assert stale.status_code == 409
    detail = await client.get(f"/api/v1/projects/{project_id}/reports/{report_id}")
    assert detail.json()["data"]["pages"][0]["blocks"][0]["version"] == 2
    assert detail.json()["data"]["pages"][0]["blocks"][0]["content"]["rows"] == [
        {"region": "华东", "revenue": 120}
    ]


@pytest.mark.asyncio
async def test_refresh_rechecks_saved_method_after_releasing_read_transaction(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    project_id = project.id
    report_id = report.id
    page_id = report.pages[0].id
    block_id = report.pages[0].blocks[0].id
    old_content = dict(report.pages[0].blocks[0].content)
    changed_playbook = _refresh_playbook(
        playbook_id=playbook.id,
        query=playbook.query,
        limit=50,
    )

    async def load_context(*_args, **_kwargs):
        return SimpleNamespace(
            sources=[],
            project_dir=Path("/tmp"),
            connection_configs={},
        )

    async def run_after_governance_change(*_args, **_kwargs):
        assert not db_session.in_transaction()
        current_project = await db_session.get(Project, project_id)
        assert current_project is not None
        current_project.extra_data = {
            "analysis_playbooks": [changed_playbook.model_dump(mode="json")]
        }
        await db_session.commit()
        return _fake_refresh_result(
            playbook,
            [{"region": "华东", "revenue": 120}],
        )

    monkeypatch.setattr(reports_api, "load_project_context", load_context)
    monkeypatch.setattr(
        reports_api,
        "run_analysis_playbook",
        run_after_governance_change,
    )
    monkeypatch.setattr(
        reports_api,
        "validate_playbook_execution_evidence",
        lambda *_args, **_kwargs: None,
    )

    rejected = await client.post(
        (
            f"/api/v1/projects/{project_id}/reports/{report_id}/pages/{page_id}"
            f"/blocks/{block_id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert rejected.status_code == 409
    current_block = await db_session.get(ReportBlock, block_id)
    current_page = await db_session.get(ReportPage, page_id)
    current_report = await db_session.get(ReportDocument, report_id)
    assert current_block is not None
    assert current_page is not None
    assert current_report is not None
    assert current_block.version == 1
    assert current_page.version == 1
    assert current_report.version == 1
    assert current_block.content == old_content
    assert "refresh_binding" not in current_block.source_ref


@pytest.mark.asyncio
async def test_full_sync_preserves_server_refresh_binding_and_ignores_client_override(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    page = report.pages[0]
    block = page.blocks[0]
    _stub_refresh_runtime(
        monkeypatch,
        _fake_refresh_result(playbook, [{"region": "华东", "revenue": 120}]),
    )
    refreshed = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )
    assert refreshed.status_code == 200, refreshed.text
    document = refreshed.json()["data"]
    refreshed_page = document["pages"][0]
    refreshed_block = refreshed_page["blocks"][0]
    trusted_binding = dict(refreshed_block["source_ref"]["refresh_binding"])
    forged_source_ref = {
        **refreshed_block["source_ref"],
        "refresh_binding": {
            "version": 1,
            "kind": "analysis_playbook",
            "playbook_id": "pb_ffffffffffffffffffff",
            "playbook_shape_hash": "f" * 64,
            "result_name": "result_9",
        },
    }

    synchronized = await client.patch(
        f"/api/v1/projects/{project.id}/reports/{report.id}",
        json={
            "expected_version": document["version"],
            "title": document["title"],
            "description": document["description"],
            "status": document["status"],
            "pages": [
                {
                    "id": refreshed_page["id"],
                    "title": refreshed_page["title"],
                    "order_index": refreshed_page["order_index"],
                    "config": refreshed_page["config"],
                    "blocks": [
                        {
                            "id": refreshed_block["id"],
                            "block_type": refreshed_block["block_type"],
                            "title": refreshed_block["title"],
                            "order_index": refreshed_block["order_index"],
                            "source_kind": refreshed_block["source_kind"],
                            "analysis_run_id": refreshed_block["analysis_run_id"],
                            "artifact_id": refreshed_block["artifact_id"],
                            "source_ref": forged_source_ref,
                            "content": {
                                **refreshed_block["content"],
                                "note": "只修改报表内容",
                            },
                            "layout": refreshed_block["layout"],
                            "config": refreshed_block["config"],
                        }
                    ],
                }
            ],
        },
    )

    assert synchronized.status_code == 200, synchronized.text
    synchronized_block = synchronized.json()["data"]["pages"][0]["blocks"][0]
    assert synchronized_block["source_ref"]["refresh_binding"] == trusted_binding
    assert (
        synchronized_block["source_ref"]["refresh_binding"]["playbook_id"]
        == playbook.id
    )
    assert synchronized_block["content"]["note"] == "只修改报表内容"


@pytest.mark.asyncio
async def test_refresh_normalizes_legacy_chart_and_removes_conflicting_row_containers(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    legacy_chart = {
        "version": 1,
        "chart_type": "bar",
        "title": "区域收入",
        "xKey": "region",
        "yKeys": ["revenue"],
        "orientation": "horizontal",
        "data": [{"region": "旧区域", "revenue": 1}],
    }
    project, playbook, _, _, report = await _seed_refreshable_report(
        db_session,
        block_type="chart",
        content={
            "chart": legacy_chart,
            "rows": legacy_chart["data"],
            "data": [{"region": "冲突数据", "revenue": 2}],
            "items": [{"region": "冲突项目", "revenue": 3}],
            "values": [{"region": "冲突值", "revenue": 4}],
        },
    )
    page = report.pages[0]
    block = page.blocks[0]
    rows = [
        {"region": "华东", "revenue": 120},
        {"region": "华南", "revenue": 80},
    ]
    _stub_refresh_runtime(monkeypatch, _fake_refresh_result(playbook, rows))

    refreshed = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert refreshed.status_code == 200, refreshed.text
    content = refreshed.json()["data"]["pages"][0]["blocks"][0]["content"]
    assert content["rows"] == rows
    assert "data" not in content
    assert "items" not in content
    assert "values" not in content
    chart = content["chart"]
    assert chart["type"] == "bar"
    assert chart["encoding"]["x"]["field"] == "region"
    assert chart["encoding"]["y"][0]["field"] == "revenue"
    assert chart["presentation"]["orientation"] == "horizontal"
    assert chart["data"] == rows
    assert chart["data_ref"]["result_name"] == "result_1"
    for legacy_key in ("chart_type", "xKey", "yKeys", "orientation"):
        assert legacy_key not in chart


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("block_type", "artifact_payload", "source_kind"),
    [
        ("text", None, "artifact"),
        ("table", None, "analysis_run"),
        (
            "chart",
            {"format": "png", "relative_path": "chart.png"},
            "artifact",
        ),
    ],
)
async def test_refresh_rejects_unsupported_block_sources(
    client: AsyncClient,
    db_session: AsyncSession,
    block_type: str,
    artifact_payload: dict | None,
    source_kind: str,
):
    project, _, _, _, report = await _seed_refreshable_report(
        db_session,
        block_type=block_type,
        artifact_payload=artifact_payload,
    )
    page = report.pages[0]
    block = page.blocks[0]
    block.source_kind = source_kind
    await db_session.commit()

    rejected = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    expected_status = 422 if source_kind != "artifact" or block_type == "text" else 409
    assert rejected.status_code == expected_status
    await db_session.refresh(block)
    assert block.version == 1


@pytest.mark.asyncio
async def test_refresh_rejects_ambiguous_lazy_binding_without_mutating_snapshot(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    duplicate = _refresh_playbook(
        playbook_id="pb_aaaaaaaaaaaaaaaaaaaa",
        query=playbook.query,
    )
    project.extra_data = {
        "analysis_playbooks": [
            playbook.model_dump(mode="json"),
            duplicate.model_dump(mode="json"),
        ]
    }
    await db_session.commit()
    page = report.pages[0]
    block = page.blocks[0]
    old_content = dict(block.content)

    rejected = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert rejected.status_code == 409
    await db_session.refresh(block)
    assert block.version == 1
    assert block.content == old_content
    assert "refresh_binding" not in block.source_ref


@pytest.mark.asyncio
async def test_refresh_rejects_bound_playbook_drift_without_mutating_snapshot(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    page = report.pages[0]
    block = page.blocks[0]
    block.source_ref = {
        **block.source_ref,
        "refresh_binding": {
            "version": 1,
            "kind": "analysis_playbook",
            "playbook_id": playbook.id,
            "playbook_shape_hash": playbook.shape_hash,
            "result_name": "result_1",
        },
    }
    changed = _refresh_playbook(
        playbook_id=playbook.id,
        query=playbook.query,
        limit=50,
    )
    project.extra_data = {
        "analysis_playbooks": [changed.model_dump(mode="json")]
    }
    await db_session.commit()
    old_content = dict(block.content)

    rejected = await client.post(
        (
            f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
            f"/blocks/{block.id}/refresh"
        ),
        json={"expected_version": 1},
    )

    assert rejected.status_code == 409
    await db_session.refresh(block)
    assert block.version == 1
    assert block.content == old_content


@pytest.mark.asyncio
async def test_refresh_runner_or_receipt_failure_leaves_old_snapshot_unchanged(
    client: AsyncClient,
    db_session: AsyncSession,
    monkeypatch: pytest.MonkeyPatch,
):
    project, playbook, _, _, report = await _seed_refreshable_report(db_session)
    page = report.pages[0]
    block = page.blocks[0]
    refresh_url = (
        f"/api/v1/projects/{project.id}/reports/{report.id}/pages/{page.id}"
        f"/blocks/{block.id}/refresh"
    )
    old_content = dict(block.content)
    old_source_ref = dict(block.source_ref)

    async def load_context(*_args, **_kwargs):
        return SimpleNamespace(
            sources=[],
            project_dir=Path("/tmp"),
            connection_configs={},
        )

    async def fail_runner(*_args, **_kwargs):
        raise AnalysisPlaybookRunnerError("internal detail")

    monkeypatch.setattr(reports_api, "load_project_context", load_context)
    monkeypatch.setattr(reports_api, "run_analysis_playbook", fail_runner)
    failed = await client.post(
        refresh_url,
        json={"expected_version": 1},
    )
    assert failed.status_code == 409
    assert "internal detail" not in failed.text

    _stub_refresh_runtime(
        monkeypatch,
        _fake_refresh_result(
            playbook,
            [{"region": "华东", "revenue": 120}],
            receipt_playbook_id="pb_ffffffffffffffffffff",
        ),
    )
    invalid_receipt = await client.post(
        refresh_url,
        json={"expected_version": 1},
    )
    assert invalid_receipt.status_code == 409

    await db_session.refresh(block)
    await db_session.refresh(page)
    await db_session.refresh(report)
    assert block.version == 1
    assert page.version == 1
    assert report.version == 1
    assert block.content == old_content
    assert block.source_ref == old_source_ref


@pytest.mark.asyncio
async def test_export_report_xlsx_is_safe_and_contains_structured_business_data(
    client: AsyncClient,
    db_session: AsyncSession,
):
    project = Project(name="Excel 导出")
    db_session.add(project)
    await db_session.commit()
    title = "七月经营/报告"
    created = await client.post(
        f"/api/v1/projects/{project.id}/reports",
        json={
            "title": title,
            "description": "供经营复盘使用",
            "pages": [
                {
                    "title": "概览",
                    "blocks": [
                        {
                            "block_type": "chart",
                            "title": "趋势/明细",
                            "order_index": 0,
                            "source_kind": "manual",
                            "content": {
                                "data": [
                                    {
                                        "月份": "=1+1",
                                        "收入": 120.5,
                                        "sql": "SELECT secret",
                                        "python": "open('/tmp/secret')",
                                        "technical_details": {"token": "secret"},
                                    },
                                    {
                                        "月份": "七月",
                                        "收入": 140,
                                        "备注": {
                                            "状态": "已确认",
                                            "sql": "SELECT nested_secret",
                                            "详情": {"python": "nested_secret()"},
                                        },
                                    },
                                ],
                                "sql": "SELECT sibling_secret",
                                "python": "print('sibling_secret')",
                                "technical_details": {"tool_history": ["secret"]},
                            },
                        },
                        {
                            "block_type": "table",
                            "title": "趋势/明细",
                            "order_index": 1,
                            "source_kind": "manual",
                            "content": {"rows": [{"门店": "南京", "订单": 12}]},
                        },
                        {
                            "block_type": "chart",
                            "title": "单列值",
                            "order_index": 2,
                            "source_kind": "manual",
                            "content": {"values": ["@SUM(A1:A2)", 2]},
                        },
                        {
                            "block_type": "text",
                            "title": "不应导出为数据页",
                            "order_index": 3,
                            "source_kind": "manual",
                            "content": {"rows": [{"sql": "SELECT hidden"}], "text": "结论"},
                        },
                    ],
                }
            ],
        },
    )
    assert created.status_code == 201, created.text
    report_id = created.json()["data"]["id"]

    response = await client.get(
        f"/api/v1/projects/{project.id}/reports/{report_id}/export.xlsx"
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        f"attachment; filename=\"report.xlsx\"; filename*=UTF-8''{quote('七月经营_报告.xlsx', safe='')}"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["报表概览", "趋势 明细", "趋势 明细 (2)", "单列值"]
    assert all(len(name) <= 31 for name in workbook.sheetnames)

    trend = workbook["趋势 明细"]
    assert [cell.value for cell in trend[1]] == ["月份", "收入", "备注"]
    assert trend["A2"].value == "'=1+1"
    assert trend["A2"].data_type == "s"
    assert trend["C3"].value == '{"状态": "已确认", "详情": {}}'
    assert workbook["单列值"]["A2"].value == "'@SUM(A1:A2)"

    all_values = {
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    assert "SELECT secret" not in all_values
    assert "SELECT nested_secret" not in all_values
    assert "open('/tmp/secret')" not in all_values
    assert "nested_secret()" not in all_values
    assert "technical_details" not in all_values
