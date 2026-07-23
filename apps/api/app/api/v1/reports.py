"""Editable report documents, pages and blocks."""

from __future__ import annotations

import json
import math
import re
from collections.abc import Sequence
from copy import deepcopy
from datetime import date, datetime, time
from io import BytesIO
from typing import Any
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db import get_db
from app.db.tables import (
    AnalysisRun,
    ArtifactRecord,
    Project,
    ReportBlock,
    ReportDocument,
    ReportPage,
)
from app.models.common import APIResponse
from app.models.reports import (
    ReportBlockCreate,
    ReportBlockRefreshBinding,
    ReportBlockRefreshRequest,
    ReportBlockSync,
    ReportBlockUpdate,
    ReportCreate,
    ReportDeleteResponse,
    ReportDocumentResponse,
    ReportPageCreate,
    ReportPageSync,
    ReportPageUpdate,
    ReportSummaryResponse,
    ReportUpdate,
)
from app.models.workspace import (
    AnalysisPlaybookResponse,
    ReportDraftPlanRequest,
    ReportDraftPlanResponse,
)
from app.services.analysis_checkpoint import stable_payload_hash
from app.services.analysis_playbook_runner import (
    AnalysisPlaybookRunnerError,
    AnalysisPlaybookRunResult,
    run_analysis_playbook,
)
from app.services.analyst_runtime import ChartSpec
from app.services.golden_regression import normalize_query_key
from app.services.project_context import load_project_context
from app.services.report_draft import generate_report_draft_plan
from app.services.standing_workspace import (
    StandingWorkspaceCorruptError,
    StandingWorkspaceError,
    load_analysis_playbooks,
    validate_playbook_baseline_evidence,
    validate_playbook_execution_evidence,
)

router = APIRouter()

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLSX_ROW_LIMIT = 1_048_576
_XLSX_COLUMN_LIMIT = 16_384
_XLSX_CELL_TEXT_LIMIT = 32_767
_INVALID_SHEET_TITLE = re.compile(r"[\[\]:*?/\\]")
_INVALID_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_STRUCTURED_CONTENT_KEYS = ("rows", "data", "items", "values")
_EXCLUDED_EXPORT_FIELDS = {"sql", "python", "technical_details", "technicaldetails"}
_BLOCK_TYPE_LABELS = {
    "metric": "指标",
    "chart": "图表",
    "table": "表格",
    "text": "文字",
    "evidence": "依据",
    "filter": "筛选",
}
_REPORT_STATUS_LABELS = {"draft": "草稿", "published": "已发布", "archived": "已归档"}
_REPORT_TABLE_SAMPLE_LIMIT = 200
_REPORT_CHART_ROW_LIMIT = 1000


def _excel_safe_text(value: str) -> str:
    """Return text that is valid XML and cannot be interpreted as an Excel formula."""

    text = _INVALID_XML_CHARACTERS.sub("", value)
    first_visible = text.lstrip()[:1]
    if text[:1] in {"\t", "\r", "\n"} or first_visible in _FORMULA_PREFIXES:
        text = f"'{text}"
    if len(text) > _XLSX_CELL_TEXT_LIMIT:
        text = f"{text[: _XLSX_CELL_TEXT_LIMIT - 1]}…"
    return text


def _excel_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, bool):
        return value
    if isinstance(value, int):
        # Excel only preserves 15 significant digits in numeric cells.
        return value if abs(value) <= 999_999_999_999_999 else _excel_safe_text(str(value))
    if isinstance(value, float):
        return value if math.isfinite(value) else _excel_safe_text(str(value))
    if isinstance(value, (date, datetime, time)):
        return _excel_safe_text(value.isoformat())
    if isinstance(value, str):
        return _excel_safe_text(value)
    if isinstance(value, (dict, list, tuple)):
        return _excel_safe_text(
            json.dumps(
                _sanitize_export_value(value),
                ensure_ascii=False,
                sort_keys=True,
                default=str,
            )
        )
    return _excel_safe_text(str(value))


def _is_exportable_field(name: str) -> bool:
    return name.strip().casefold() not in _EXCLUDED_EXPORT_FIELDS


def _sanitize_export_value(value: Any) -> Any:
    """Remove technical fields recursively before serializing nested cell values."""

    if isinstance(value, dict):
        return {
            str(key): _sanitize_export_value(item)
            for key, item in value.items()
            if _is_exportable_field(str(key))
        }
    if isinstance(value, (list, tuple)):
        return [_sanitize_export_value(item) for item in value]
    return value


def _structured_block_rows(
    block: ReportBlock,
) -> tuple[list[str], list[list[Any]]] | None:
    """Normalize the structured content formats understood by the report UI."""

    if block.block_type not in {"chart", "table"}:
        return None
    content = block.content if isinstance(block.content, dict) else {}
    for key in _STRUCTURED_CONTENT_KEYS:
        candidate = content.get(key)
        if not isinstance(candidate, list) or not candidate:
            continue

        if all(isinstance(item, dict) for item in candidate):
            columns: list[str] = []
            seen: set[str] = set()
            for item in candidate:
                for raw_name in item:
                    name = str(raw_name)
                    if not _is_exportable_field(name) or name in seen:
                        continue
                    seen.add(name)
                    columns.append(name)
            if not columns:
                continue
            return columns, [[item.get(column) for column in columns] for item in candidate]

        if all(isinstance(item, (list, tuple)) for item in candidate):
            width = max((len(item) for item in candidate), default=0)
            if width == 0:
                continue
            if width > _XLSX_COLUMN_LIMIT:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="报表数据列数超过 Excel 支持范围",
                )
            configured_columns = content.get("columns")
            raw_columns = list(configured_columns) if isinstance(configured_columns, list) else []
            indexed_columns = [
                str(raw_columns[index]) if index < len(raw_columns) else f"列 {index + 1}"
                for index in range(width)
            ]
            kept_indexes = [
                index
                for index, name in enumerate(indexed_columns)
                if _is_exportable_field(name)
            ]
            if not kept_indexes:
                continue
            return (
                [indexed_columns[index] for index in kept_indexes],
                [
                    [item[index] if index < len(item) else None for index in kept_indexes]
                    for item in candidate
                ],
            )

        return ["值"], [[item] for item in candidate]
    return None


def _unique_sheet_title(raw_title: str, used: set[str]) -> str:
    clean_title = _INVALID_XML_CHARACTERS.sub("", raw_title)
    base = _INVALID_SHEET_TITLE.sub(" ", clean_title).strip(" '\t\r\n") or "数据"
    base = base[:31]
    candidate = base
    suffix_index = 2
    while candidate.casefold() in used:
        suffix = f" ({suffix_index})"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        suffix_index += 1
    used.add(candidate.casefold())
    return candidate


def _style_header_row(sheet: Any) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F5E4B")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def _fit_sheet_columns(sheet: Any) -> None:
    for column_cells in sheet.iter_cols():
        width = min(
            40,
            max(
                10,
                max(
                    (len(str(cell.value)) for cell in column_cells if cell.value is not None),
                    default=0,
                )
                + 2,
            ),
        )
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _safe_xlsx_filename(title: str) -> str:
    safe_title = re.sub(r"[\\/\r\n\x00-\x1f]", "_", title).strip(" .") or "报告"
    return f"{safe_title[:120]}.xlsx"


async def _require_project(db: AsyncSession, project_id: UUID) -> Project:
    project = await db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    return project


async def _load_report(
    db: AsyncSession,
    project_id: UUID,
    report_id: UUID,
) -> ReportDocument:
    result = await db.execute(
        select(ReportDocument)
        .where(
            ReportDocument.id == report_id,
            ReportDocument.project_id == project_id,
        )
        .options(selectinload(ReportDocument.pages).selectinload(ReportPage.blocks))
        .execution_options(populate_existing=True)
    )
    report = result.scalar_one_or_none()
    if report is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="报表不存在")
    return report


async def _response_report(
    db: AsyncSession,
    project_id: UUID,
    report_id: UUID,
) -> ReportDocumentResponse:
    report = await _load_report(db, project_id, report_id)
    return ReportDocumentResponse.model_validate(report)


async def _validate_source_reference(
    db: AsyncSession,
    *,
    project_id: UUID,
    source_kind: str,
    analysis_run_id: UUID | None,
    artifact_id: UUID | None,
) -> tuple[UUID | None, UUID | None, dict[str, Any]]:
    """Validate live provenance and capture a durable, immutable source snapshot."""

    if source_kind == "manual":
        if analysis_run_id is not None or artifact_id is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="人工区块不能冒充调查或产物来源",
            )
        return None, None, {}

    if source_kind == "analysis_run":
        if analysis_run_id is None or artifact_id is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="调查区块必须只引用一个 analysis_run_id",
            )
        run = await db.get(AnalysisRun, analysis_run_id)
        if run is None or run.project_id != project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="调查引用不属于当前项目",
            )
        return (
            run.id,
            None,
            {
                "source_kind": "analysis_run",
                "analysis_run_id": str(run.id),
                "captured_at": run.updated_at.isoformat(),
                "snapshot": {
                    "query": run.query,
                    "state": run.state,
                    "stage": run.stage,
                    "report": run.report,
                },
            },
        )

    if source_kind == "artifact":
        if artifact_id is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="产物区块必须引用 artifact_id",
            )
        artifact = await db.get(ArtifactRecord, artifact_id)
        if artifact is None or artifact.project_id != project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="产物引用不属于当前项目",
            )
        if analysis_run_id is not None and analysis_run_id != artifact.analysis_run_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="产物与调查引用不一致",
            )
        run = await db.get(AnalysisRun, artifact.analysis_run_id)
        return (
            artifact.analysis_run_id,
            artifact.id,
            {
                "source_kind": "artifact",
                "artifact_id": str(artifact.id),
                "analysis_run_id": str(artifact.analysis_run_id),
                "captured_at": artifact.updated_at.isoformat(),
                "snapshot": {
                    "artifact_kind": artifact.kind,
                    "title": artifact.title,
                    "payload": artifact.payload,
                    "analysis_query": run.query if run is not None else None,
                },
            },
        )

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="未知的区块来源",
    )


def _block_values(
    payload: ReportBlockCreate | ReportBlockSync,
    *,
    analysis_run_id: UUID | None,
    artifact_id: UUID | None,
    source_ref: dict[str, Any],
) -> dict[str, Any]:
    return {
        "block_type": payload.block_type,
        "title": payload.title,
        "order_index": payload.order_index,
        "source_kind": payload.source_kind,
        "analysis_run_id": analysis_run_id,
        "artifact_id": artifact_id,
        "source_ref": source_ref,
        "content": payload.content,
        "layout": payload.layout,
        "config": payload.config,
    }


async def _validated_block_values(
    db: AsyncSession,
    project_id: UUID,
    payload: ReportBlockCreate | ReportBlockSync,
    *,
    existing: ReportBlock | None = None,
) -> dict[str, Any]:
    if (
        existing is not None
        and existing.source_kind == payload.source_kind
        and payload.analysis_run_id is None
        and payload.artifact_id is None
        and (existing.source_kind == "manual" or existing.source_ref)
    ):
        return _block_values(
            payload,
            analysis_run_id=existing.analysis_run_id,
            artifact_id=existing.artifact_id,
            source_ref=existing.source_ref,
        )

    analysis_run_id, artifact_id, source_ref = await _validate_source_reference(
        db,
        project_id=project_id,
        source_kind=payload.source_kind,
        analysis_run_id=payload.analysis_run_id,
        artifact_id=payload.artifact_id,
    )
    if (
        existing is not None
        and existing.source_kind == payload.source_kind
        and existing.analysis_run_id == analysis_run_id
        and existing.artifact_id == artifact_id
    ):
        raw_refresh_binding = (existing.source_ref or {}).get("refresh_binding")
        if raw_refresh_binding is not None:
            try:
                trusted_refresh_binding = ReportBlockRefreshBinding.model_validate(
                    raw_refresh_binding
                )
            except ValueError:
                trusted_refresh_binding = None
            if trusted_refresh_binding is not None:
                source_ref = {
                    **source_ref,
                    "refresh_binding": trusted_refresh_binding.model_dump(mode="json"),
                }
    return _block_values(
        payload,
        analysis_run_id=analysis_run_id,
        artifact_id=artifact_id,
        source_ref=source_ref,
    )


async def _append_page(
    db: AsyncSession,
    *,
    report: ReportDocument,
    project_id: UUID,
    payload: ReportPageCreate,
    validated_blocks: Sequence[dict[str, Any]] | None = None,
) -> ReportPage:
    block_values = list(validated_blocks) if validated_blocks is not None else [
        await _validated_block_values(db, project_id, block) for block in payload.blocks
    ]
    page = ReportPage(
        report_id=report.id,
        title=payload.title,
        order_index=payload.order_index,
        config=payload.config,
        blocks=[],
    )
    report.pages.append(page)
    for values in block_values:
        page.blocks.append(ReportBlock(**values))
    return page


async def _validate_sync_tree(
    db: AsyncSession,
    *,
    report: ReportDocument,
    project_id: UUID,
    pages: Sequence[ReportPageSync],
) -> dict[tuple[int, int], dict[str, Any]]:
    existing_pages = {page.id: page for page in report.pages}
    seen_page_ids: set[UUID] = set()
    seen_block_ids: set[UUID] = set()
    validated: dict[tuple[int, int], dict[str, Any]] = {}

    for page_index, page_payload in enumerate(pages):
        existing_page = None
        if page_payload.id is not None:
            if page_payload.id in seen_page_ids:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="同一页面不能在报表中出现两次",
                )
            seen_page_ids.add(page_payload.id)
            existing_page = existing_pages.get(page_payload.id)
            if existing_page is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="页面不属于当前报表",
                )
        existing_blocks = {
            block.id: block for block in existing_page.blocks
        } if existing_page is not None else {}

        for block_index, block_payload in enumerate(page_payload.blocks):
            if block_payload.id is not None:
                if block_payload.id in seen_block_ids:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="同一区块不能在报表中出现两次",
                    )
                seen_block_ids.add(block_payload.id)
                if block_payload.id not in existing_blocks:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="区块不属于指定页面",
                    )
            existing_block = (
                existing_blocks.get(block_payload.id)
                if block_payload.id is not None
                else None
            )
            validated[(page_index, block_index)] = await _validated_block_values(
                db,
                project_id,
                block_payload,
                existing=existing_block,
            )
    return validated


async def _sync_pages(
    db: AsyncSession,
    *,
    report: ReportDocument,
    project_id: UUID,
    pages: Sequence[ReportPageSync],
    validated_blocks: dict[tuple[int, int], dict[str, Any]] | None = None,
) -> None:
    if validated_blocks is None:
        validated_blocks = await _validate_sync_tree(
            db,
            report=report,
            project_id=project_id,
            pages=pages,
        )
    existing_pages = {page.id: page for page in report.pages}
    kept_page_ids = {page.id for page in pages if page.id is not None}

    for old_page in list(report.pages):
        if old_page.id not in kept_page_ids:
            report.pages.remove(old_page)

    for page_index, page_payload in enumerate(pages):
        if page_payload.id is None:
            page = ReportPage(report_id=report.id, blocks=[])
            report.pages.append(page)
        else:
            page = existing_pages[page_payload.id]
            page.version += 1
        page.title = page_payload.title
        page.order_index = page_payload.order_index
        page.config = page_payload.config

        existing_blocks = {block.id: block for block in page.blocks}
        kept_block_ids = {block.id for block in page_payload.blocks if block.id is not None}
        for old_block in list(page.blocks):
            if old_block.id not in kept_block_ids:
                page.blocks.remove(old_block)

        for block_index, block_payload in enumerate(page_payload.blocks):
            if block_payload.id is None:
                block = ReportBlock(page_id=page.id)
                page.blocks.append(block)
            else:
                block = existing_blocks[block_payload.id]
                block.version += 1
            for field_name, value in validated_blocks[(page_index, block_index)].items():
                setattr(block, field_name, value)


def _assert_expected_version(current: int, expected: int) -> None:
    if current != expected:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="报表已被更新，请重新载入后再保存",
        )


async def _cas_entity_version(
    db: AsyncSession,
    model: Any,
    entity_id: UUID,
    expected: int,
    *,
    detail: str,
) -> None:
    """Atomically reserve one entity version; only one stale writer can win."""

    result = await db.execute(
        update(model)
        .where(model.id == entity_id, model.version == expected)
        .values(version=expected + 1)
        .execution_options(synchronize_session="fetch")
    )
    if result.rowcount != 1:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=detail,
        )


async def _bump_entity_version(
    db: AsyncSession,
    model: Any,
    entity_id: UUID,
    *,
    detail: str,
) -> None:
    """Atomically mark a parent as changed after a nested mutation."""

    result = await db.execute(
        update(model)
        .where(model.id == entity_id)
        .values(version=model.version + 1)
        .execution_options(synchronize_session="fetch")
    )
    if result.rowcount != 1:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=detail)


async def _claim_report_version(
    db: AsyncSession,
    report: ReportDocument,
    expected: int,
) -> None:
    await _cas_entity_version(
        db,
        ReportDocument,
        report.id,
        expected,
        detail="报表已被更新，请重新载入后再保存",
    )


@router.get(
    "/projects/{project_id}/reports",
    response_model=APIResponse[list[ReportSummaryResponse]],
)
async def list_reports(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    """List editable reports without loading their full document trees."""

    await _require_project(db, project_id)
    page_count = (
        select(func.count(ReportPage.id))
        .where(ReportPage.report_id == ReportDocument.id)
        .correlate(ReportDocument)
        .scalar_subquery()
    )
    block_count = (
        select(func.count(ReportBlock.id))
        .join(ReportPage, ReportPage.id == ReportBlock.page_id)
        .where(ReportPage.report_id == ReportDocument.id)
        .correlate(ReportDocument)
        .scalar_subquery()
    )
    result = await db.execute(
        select(
            ReportDocument,
            page_count.label("page_count"),
            block_count.label("block_count"),
        )
        .where(ReportDocument.project_id == project_id)
        .order_by(ReportDocument.updated_at.desc())
    )
    summaries = [
        ReportSummaryResponse(
            id=report.id,
            project_id=report.project_id,
            title=report.title,
            description=report.description,
            status=report.status,
            version=report.version,
            page_count=int(pages or 0),
            block_count=int(blocks or 0),
            created_at=report.created_at,
            updated_at=report.updated_at,
        )
        for report, pages, blocks in result.all()
    ]
    return APIResponse.ok(data=summaries)


@router.post(
    "/projects/{project_id}/reports",
    response_model=APIResponse[ReportDocumentResponse],
    status_code=status.HTTP_201_CREATED,
)
async def create_report(
    project_id: UUID,
    payload: ReportCreate,
    db: AsyncSession = Depends(get_db),
):
    """Create a report, optionally from a complete initial page/block tree."""

    await _require_project(db, project_id)
    pages = payload.pages or [ReportPageCreate()]
    validated_pages = [
        [await _validated_block_values(db, project_id, block) for block in page.blocks]
        for page in pages
    ]
    report = ReportDocument(
        project_id=project_id,
        title=payload.title,
        description=payload.description,
        status=payload.status,
        extra_data=payload.extra_data,
        pages=[],
    )
    db.add(report)
    await db.flush()
    for page_payload, validated_blocks in zip(pages, validated_pages, strict=True):
        await _append_page(
            db,
            report=report,
            project_id=project_id,
            payload=page_payload,
            validated_blocks=validated_blocks,
        )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report.id))


@router.post(
    "/projects/{project_id}/reports/draft-from-analysis",
    response_model=APIResponse[ReportDraftPlanResponse],
)
async def draft_report_from_analysis(
    project_id: UUID,
    payload: ReportDraftPlanRequest,
    db: AsyncSession = Depends(get_db),
):
    """Use the configured model to plan a smart draft from an analysis run."""

    await _require_project(db, project_id)
    plan = await generate_report_draft_plan(
        db,
        project_id=project_id,
        run_id=payload.analysis_run_id,
        language=payload.language,
        current_report=payload.current_report,
    )
    if plan is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "Smart report planning is temporarily unavailable. Please try again later "
                "or select content manually."
                if payload.language == "en"
                else "智能整理暂不可用，请稍后再试或手动选择内容"
            ),
        )
    response = ReportDraftPlanResponse(
        title=plan.title,
        description=plan.description,
        overview_text=plan.overview_text,
        sections=[section.model_dump() for section in plan.sections],
        selected_overview=plan.selected_overview,
        selected_detail=plan.selected_detail,
        selected_evidence=plan.selected_evidence,
        highlights=plan.highlights,
        generated_by="ai",
    )
    return APIResponse.ok(data=response)


@router.get(
    "/projects/{project_id}/reports/{report_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def get_report(
    project_id: UUID,
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    await _require_project(db, project_id)
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


@router.get("/projects/{project_id}/reports/{report_id}/export.xlsx")
async def export_report_xlsx(
    project_id: UUID,
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Export the persisted business report without its technical provenance."""

    await _require_project(db, project_id)
    report = await _load_report(db, project_id, report_id)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "报表概览"
    workbook.properties.title = _excel_safe_text(report.title)
    workbook.properties.creator = "ReceiptBI"

    overview.append(["项目", "内容"])
    overview.append(["报告标题", _excel_cell_value(report.title)])
    overview.append(["说明", _excel_cell_value(report.description or "")])
    overview.append(["状态", _REPORT_STATUS_LABELS.get(report.status, report.status)])
    overview.append(["版本", report.version])
    overview.append(["创建时间", report.created_at.isoformat()])
    overview.append(["更新时间", report.updated_at.isoformat()])
    overview.append([])
    overview.append(["页面", "区块", "类型", "数据工作表"])

    used_sheet_titles = {overview.title.casefold()}
    for page in sorted(report.pages, key=lambda item: (item.order_index, str(item.id))):
        for block in sorted(page.blocks, key=lambda item: (item.order_index, str(item.id))):
            structured = _structured_block_rows(block)
            data_sheet_title = ""
            if structured is not None:
                columns, rows = structured
                if len(columns) > _XLSX_COLUMN_LIMIT:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="报表数据列数超过 Excel 支持范围",
                    )
                if len(rows) + 1 > _XLSX_ROW_LIMIT:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail="报表数据行数超过 Excel 支持范围",
                    )
                raw_sheet_title = block.title or f"{page.title}数据"
                data_sheet_title = _unique_sheet_title(raw_sheet_title, used_sheet_titles)
                sheet = workbook.create_sheet(title=data_sheet_title)
                sheet.append([_excel_safe_text(column) for column in columns])
                for row in rows:
                    sheet.append([_excel_cell_value(value) for value in row])
                _style_header_row(sheet)
                _fit_sheet_columns(sheet)
                sheet.auto_filter.ref = sheet.dimensions

            overview.append(
                [
                    _excel_cell_value(page.title),
                    _excel_cell_value(block.title or "未命名区块"),
                    _BLOCK_TYPE_LABELS.get(block.block_type, block.block_type),
                    _excel_cell_value(data_sheet_title),
                ]
            )

    _style_header_row(overview)
    _fit_sheet_columns(overview)

    output = BytesIO()
    workbook.save(output)
    filename = _safe_xlsx_filename(report.title)
    encoded_filename = quote(filename, safe="")
    return Response(
        content=output.getvalue(),
        media_type=_XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"report.xlsx\"; filename*=UTF-8''{encoded_filename}"
            ),
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


@router.patch(
    "/projects/{project_id}/reports/{report_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def update_report(
    project_id: UUID,
    report_id: UUID,
    payload: ReportUpdate,
    db: AsyncSession = Depends(get_db),
):
    """Update metadata or transactionally synchronize the complete document tree."""

    await _require_project(db, project_id)
    report = await _load_report(db, project_id, report_id)
    _assert_expected_version(report.version, payload.expected_version)
    fields = payload.model_fields_set - {"expected_version"}
    if not fields:
        return APIResponse.ok(data=ReportDocumentResponse.model_validate(report))

    validated_blocks = None
    if payload.pages is not None:
        validated_blocks = await _validate_sync_tree(
            db,
            report=report,
            project_id=project_id,
            pages=payload.pages,
        )
    await _claim_report_version(db, report, payload.expected_version)
    if payload.pages is not None:
        await _sync_pages(
            db,
            report=report,
            project_id=project_id,
            pages=payload.pages,
            validated_blocks=validated_blocks,
        )
    for field_name in ("title", "description", "status", "extra_data"):
        if field_name in payload.model_fields_set:
            setattr(report, field_name, getattr(payload, field_name))
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report.id))


@router.delete(
    "/projects/{project_id}/reports/{report_id}",
    response_model=APIResponse[ReportDeleteResponse],
)
async def delete_report(
    project_id: UUID,
    report_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    await _require_project(db, project_id)
    report = await _load_report(db, project_id, report_id)
    await db.delete(report)
    await db.commit()
    return APIResponse.ok(data=ReportDeleteResponse(id=report_id))


@router.post(
    "/projects/{project_id}/reports/{report_id}/pages",
    response_model=APIResponse[ReportDocumentResponse],
    status_code=status.HTTP_201_CREATED,
)
async def create_report_page(
    project_id: UUID,
    report_id: UUID,
    payload: ReportPageCreate,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    await _append_page(
        db,
        report=report,
        project_id=project_id,
        payload=payload,
    )
    await _bump_entity_version(
        db,
        ReportDocument,
        report.id,
        detail="报表已被删除或更新，请重新载入",
    )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


def _page_from_report(report: ReportDocument, page_id: UUID) -> ReportPage:
    page = next((item for item in report.pages if item.id == page_id), None)
    if page is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="页面不存在")
    return page


@router.patch(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def update_report_page(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    payload: ReportPageUpdate,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    page = _page_from_report(report, page_id)
    _assert_expected_version(page.version, payload.expected_version)
    fields = payload.model_fields_set - {"expected_version"}
    if fields:
        await _cas_entity_version(
            db,
            ReportPage,
            page.id,
            payload.expected_version,
            detail="页面已被更新，请重新载入后再保存",
        )
        for field_name in ("title", "order_index", "config"):
            if field_name in payload.model_fields_set:
                setattr(page, field_name, getattr(payload, field_name))
        await _bump_entity_version(
            db,
            ReportDocument,
            report.id,
            detail="报表已被删除，请重新载入",
        )
        await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


@router.delete(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def delete_report_page(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    page = _page_from_report(report, page_id)
    await db.delete(page)
    await _bump_entity_version(
        db,
        ReportDocument,
        report.id,
        detail="报表已被删除，请重新载入",
    )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


@router.post(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}/blocks",
    response_model=APIResponse[ReportDocumentResponse],
    status_code=status.HTTP_201_CREATED,
)
async def create_report_block(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    payload: ReportBlockCreate,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    page = _page_from_report(report, page_id)
    values = await _validated_block_values(db, project_id, payload)
    page.blocks.append(ReportBlock(page_id=page.id, **values))
    await _bump_entity_version(
        db,
        ReportPage,
        page.id,
        detail="页面已被删除，请重新载入",
    )
    await _bump_entity_version(
        db,
        ReportDocument,
        report.id,
        detail="报表已被删除，请重新载入",
    )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


def _block_from_page(page: ReportPage, block_id: UUID) -> ReportBlock:
    block = next((item for item in page.blocks if item.id == block_id), None)
    if block is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="区块不存在")
    return block


def _final_run_validation(
    run: AnalysisRun,
) -> tuple[list[dict[str, Any]], dict[str, Any], str]:
    tool_history = [
        dict(item)
        for item in (run.checkpoint or {}).get("tool_history") or []
        if isinstance(item, dict)
    ]
    validations = [
        (index, item)
        for index, item in enumerate(tool_history)
        if item.get("kind") == "validation" and isinstance(item.get("profile"), dict)
    ]
    if not validations:
        raise StandingWorkspaceError("调查没有可用的最终结果")
    validation_index, validation = validations[-1]
    result_name = str(validation.get("result_name") or "").strip()
    data_steps = [
        (index, item)
        for index, item in enumerate(tool_history)
        if item.get("kind")
        in {
            "structured_query",
            "sql",
            "file_sql",
            "join",
            "aggregate",
            "business_rule_application",
        }
        and str(item.get("result_name") or "").strip()
    ]
    if (
        not result_name
        or not data_steps
        or validation_index < data_steps[-1][0]
        or result_name != str(data_steps[-1][1].get("result_name") or "").strip()
        or (validation.get("profile") or {}).get("truncated") is not False
    ):
        raise StandingWorkspaceError("调查没有可用的最终结果")
    return tool_history, validation, result_name


def _refreshable_artifact(block: ReportBlock, artifact: ArtifactRecord) -> bool:
    if block.source_kind != "artifact":
        return False
    if block.block_type == "table":
        return artifact.kind == "table"
    if block.block_type != "chart" or artifact.kind != "chart":
        return False
    payload = artifact.payload if isinstance(artifact.payload, dict) else {}
    if str(payload.get("format") or "").casefold() == "png":
        return False
    return isinstance(payload.get("chart"), dict)


def _system_playbook_matches(
    playbook: AnalysisPlaybookResponse,
    *,
    query_key: str,
    result_name: str,
) -> bool:
    return bool(
        playbook.schema_version == 3
        and playbook.execution_mode == "system_structured_query"
        and normalize_query_key(playbook.query) == query_key
        and playbook.validation.input_result == result_name
    )


async def _resolve_refresh_binding(
    db: AsyncSession,
    *,
    project: Project,
    block: ReportBlock,
) -> tuple[AnalysisPlaybookResponse, ReportBlockRefreshBinding]:
    try:
        playbooks = load_analysis_playbooks(project)
    except (StandingWorkspaceCorruptError, StandingWorkspaceError) as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块的刷新来源已不可用，请重新添加",
        ) from exc

    raw_binding = (block.source_ref or {}).get("refresh_binding")
    if raw_binding is not None:
        try:
            binding = ReportBlockRefreshBinding.model_validate(raw_binding)
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="这个区块的刷新来源已变化，请重新添加",
            ) from exc
        matches = [item for item in playbooks if item.id == binding.playbook_id]
        if len(matches) != 1:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="这个区块的刷新来源已变化，请重新添加",
            )
        playbook = matches[0]
        if (
            playbook.shape_hash != binding.playbook_shape_hash
            or playbook.validation.input_result != binding.result_name
            or not _system_playbook_matches(
                playbook,
                query_key=normalize_query_key(playbook.query),
                result_name=binding.result_name,
            )
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="这个区块的刷新来源已变化，请重新添加",
            )
        return playbook, binding

    if block.artifact_id is None or block.analysis_run_id is None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块没有可用的刷新来源，请重新添加",
        )
    artifact = await db.get(ArtifactRecord, block.artifact_id)
    run = await db.get(AnalysisRun, block.analysis_run_id)
    if (
        artifact is None
        or run is None
        or artifact.project_id != project.id
        or artifact.analysis_run_id != run.id
        or run.project_id != project.id
        or run.state != "completed"
        or not _refreshable_artifact(block, artifact)
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块没有可用的刷新来源，请重新添加",
        )

    try:
        tool_history, validation, result_name = _final_run_validation(run)
    except StandingWorkspaceError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块没有可用的刷新来源，请重新添加",
        ) from exc
    if str((artifact.technical_details or {}).get("result_name") or "") != result_name:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块没有可用的刷新来源，请重新添加",
        )
    query_key = normalize_query_key(run.query)
    matches = [
        item
        for item in playbooks
        if _system_playbook_matches(
            item,
            query_key=query_key,
            result_name=result_name,
        )
    ]
    if len(matches) != 1:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="无法确认这个区块的唯一刷新来源，请重新添加",
        )
    playbook = matches[0]
    try:
        validate_playbook_baseline_evidence(playbook, tool_history, validation)
    except StandingWorkspaceError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="这个区块没有可用的刷新来源，请重新添加",
        ) from exc
    return playbook, ReportBlockRefreshBinding(
        playbook_id=playbook.id,
        playbook_shape_hash=playbook.shape_hash,
        result_name=result_name,
    )


def _validate_refresh_result(
    playbook: AnalysisPlaybookResponse,
    result: AnalysisPlaybookRunResult,
    binding: ReportBlockRefreshBinding,
) -> None:
    receipt = result.receipt
    if (
        result.result_name != binding.result_name
        or result.validated_results != {binding.result_name}
        or receipt.playbook_id != binding.playbook_id
        or receipt.playbook_shape_hash != binding.playbook_shape_hash
        or receipt.result_name != binding.result_name
        or receipt.row_count != len(result.rows)
        or receipt.truncated is not False
        or not all(isinstance(row, dict) for row in result.rows)
    ):
        raise StandingWorkspaceError("刷新结果没有通过校验")
    validate_playbook_execution_evidence(
        playbook,
        result.tool_history,
        result.validation,
    )


def _refresh_playbook_contract(playbook: AnalysisPlaybookResponse) -> dict[str, Any]:
    """Return only the immutable execution contract used by report refresh."""

    return {
        "schema_version": playbook.schema_version,
        "execution_mode": playbook.execution_mode,
        "id": playbook.id,
        "binding_policy": playbook.binding_policy,
        "requires_revalidation": playbook.requires_revalidation,
        "source_roles": [
            item.model_dump(mode="json") for item in playbook.source_roles
        ],
        "confirmed_knowledge_keys": list(playbook.confirmed_knowledge_keys),
        "relationship_keys": list(playbook.relationship_keys),
        "steps": [item.model_dump(mode="json") for item in playbook.steps],
        "validation": playbook.validation.model_dump(mode="json"),
        "shape_hash": playbook.shape_hash,
    }


async def _lock_project_for_report_refresh(
    db: AsyncSession,
    project_id: UUID,
) -> Project:
    """Open the short write transaction and serialize catalog governance changes."""

    bind = db.get_bind()
    if bind is not None and bind.dialect.name == "sqlite":
        # SQLite ignores SELECT FOR UPDATE. BEGIN IMMEDIATE gives this short
        # write phase the same serialization boundary before the catalog is read.
        await db.execute(text("BEGIN IMMEDIATE"))
        statement = select(Project).where(Project.id == project_id)
    else:
        statement = (
            select(Project)
            .where(Project.id == project_id)
            .with_for_update()
        )
    result = await db.execute(statement.execution_options(populate_existing=True))
    project = result.scalar_one_or_none()
    if project is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    return project


def _without_temporary_filters(content: dict[str, Any]) -> dict[str, Any]:
    refreshed = {
        key: value
        for key, value in content.items()
        if not str(key).startswith("_filter_")
    }
    for key in ("data", "items", "values"):
        refreshed.pop(key, None)
    return refreshed


def _refreshed_block_content(
    block: ReportBlock,
    *,
    rows: list[dict[str, Any]],
    result_name: str,
    columns: list[str],
) -> dict[str, Any]:
    content = _without_temporary_filters(dict(block.content or {}))
    if block.block_type == "table":
        content.update(
            {
                "rows": rows[:_REPORT_TABLE_SAMPLE_LIMIT],
                "rows_count": len(rows),
                "sampled": len(rows) > _REPORT_TABLE_SAMPLE_LIMIT,
                "columns": columns,
            }
        )
        return content

    if len(rows) > _REPORT_CHART_ROW_LIMIT:
        raise StandingWorkspaceError("当前结果不适合直接刷新图表")
    raw_chart = content.get("chart")
    if not isinstance(raw_chart, dict):
        raise StandingWorkspaceError("图表内容不可刷新")
    try:
        validated_chart = ChartSpec.model_validate(raw_chart)
    except ValueError as exc:
        raise StandingWorkspaceError("图表内容不可刷新") from exc
    chart = validated_chart.model_dump(mode="json", exclude_none=True)
    chart["data"] = rows
    chart["data_ref"] = {
        "result_name": result_name,
        "result_hash": stable_payload_hash(rows),
    }
    content["chart"] = chart
    content["rows"] = rows
    return content


@router.post(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}/blocks/{block_id}/refresh",
    response_model=APIResponse[ReportDocumentResponse],
)
async def refresh_report_block(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    block_id: UUID,
    payload: ReportBlockRefreshRequest,
    db: AsyncSession = Depends(get_db),
):
    """Refresh a table or structured chart from its governed saved method."""

    try:
        project = await _require_project(db, project_id)
        report = await _load_report(db, project_id, report_id)
        page = _page_from_report(report, page_id)
        block = _block_from_page(page, block_id)
        _assert_expected_version(block.version, payload.expected_version)
        if block.block_type not in {"table", "chart"} or block.source_kind != "artifact":
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="这个区块暂不支持刷新",
            )

        playbook, binding = await _resolve_refresh_binding(
            db,
            project=project,
            block=block,
        )
        context = await load_project_context(db, project_id)
        execution_playbook = playbook.model_copy(deep=True)
        execution_binding = binding.model_copy(deep=True)
        execution_sources = deepcopy(context.sources)
        execution_project_dir = context.project_dir
        execution_connection_configs = deepcopy(context.connection_configs)
        await db.rollback()

        try:
            result = await run_analysis_playbook(
                execution_playbook,
                sources=execution_sources,
                project_dir=execution_project_dir,
                connection_configs=execution_connection_configs,
            )
            _validate_refresh_result(
                execution_playbook,
                result,
                execution_binding,
            )
        except (AnalysisPlaybookRunnerError, StandingWorkspaceError, ValueError) as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="当前数据暂时无法刷新这个区块",
            ) from exc

        current_project = await _lock_project_for_report_refresh(db, project_id)
        current_report = await _load_report(db, project_id, report_id)
        current_page = _page_from_report(current_report, page_id)
        current_block = _block_from_page(current_page, block_id)
        _assert_expected_version(current_block.version, payload.expected_version)
        if (
            current_block.block_type not in {"table", "chart"}
            or current_block.source_kind != "artifact"
        ):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="这个区块暂不支持刷新",
            )
        current_playbook, current_binding = await _resolve_refresh_binding(
            db,
            project=current_project,
            block=current_block,
        )
        if (
            current_binding != execution_binding
            or _refresh_playbook_contract(current_playbook)
            != _refresh_playbook_contract(execution_playbook)
        ):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="这个区块的刷新来源已变化，请重新刷新",
            )
        try:
            refreshed_content = _refreshed_block_content(
                current_block,
                rows=[dict(row) for row in result.rows],
                result_name=execution_binding.result_name,
                columns=list(execution_playbook.validation.columns),
            )
        except (StandingWorkspaceError, ValueError) as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="当前数据暂时无法刷新这个区块",
            ) from exc

        await _cas_entity_version(
            db,
            ReportBlock,
            current_block.id,
            payload.expected_version,
            detail="区块已被更新，请重新载入后再刷新",
        )
        current_block.content = refreshed_content
        current_block.source_ref = {
            **(current_block.source_ref or {}),
            "refresh_binding": execution_binding.model_dump(mode="json"),
        }
        await _bump_entity_version(
            db,
            ReportPage,
            current_page.id,
            detail="页面已被删除，请重新载入",
        )
        await _bump_entity_version(
            db,
            ReportDocument,
            current_report.id,
            detail="报表已被删除，请重新载入",
        )
        await db.commit()
    except Exception:
        await db.rollback()
        raise
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


@router.patch(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}/blocks/{block_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def update_report_block(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    block_id: UUID,
    payload: ReportBlockUpdate,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    page = _page_from_report(report, page_id)
    block = _block_from_page(page, block_id)
    _assert_expected_version(block.version, payload.expected_version)
    fields = payload.model_fields_set - {"expected_version"}
    if not fields:
        return APIResponse.ok(data=ReportDocumentResponse.model_validate(report))

    source_fields = {"source_kind", "analysis_run_id", "artifact_id"}
    if fields & source_fields:
        source_kind = payload.source_kind if "source_kind" in fields else block.source_kind
        if "source_kind" in fields and source_kind == "manual":
            analysis_run_id = None
            artifact_id = None
        else:
            analysis_run_id = (
                payload.analysis_run_id
                if "analysis_run_id" in fields
                else block.analysis_run_id
            )
            artifact_id = (
                payload.artifact_id if "artifact_id" in fields else block.artifact_id
            )
            if source_kind == "analysis_run":
                artifact_id = None
            elif source_kind == "artifact" and "artifact_id" in fields:
                if "analysis_run_id" not in fields:
                    analysis_run_id = None
        analysis_run_id, artifact_id, source_ref = await _validate_source_reference(
            db,
            project_id=project_id,
            source_kind=source_kind,
            analysis_run_id=analysis_run_id,
            artifact_id=artifact_id,
        )
    else:
        source_kind = block.source_kind
        analysis_run_id = block.analysis_run_id
        artifact_id = block.artifact_id
        source_ref = block.source_ref

    await _cas_entity_version(
        db,
        ReportBlock,
        block.id,
        payload.expected_version,
        detail="区块已被更新，请重新载入后再保存",
    )

    for field_name in ("block_type", "title", "order_index", "content", "layout", "config"):
        if field_name in fields:
            setattr(block, field_name, getattr(payload, field_name))
    block.source_kind = source_kind
    block.analysis_run_id = analysis_run_id
    block.artifact_id = artifact_id
    block.source_ref = source_ref
    await _bump_entity_version(
        db,
        ReportPage,
        page.id,
        detail="页面已被删除，请重新载入",
    )
    await _bump_entity_version(
        db,
        ReportDocument,
        report.id,
        detail="报表已被删除，请重新载入",
    )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))


@router.delete(
    "/projects/{project_id}/reports/{report_id}/pages/{page_id}/blocks/{block_id}",
    response_model=APIResponse[ReportDocumentResponse],
)
async def delete_report_block(
    project_id: UUID,
    report_id: UUID,
    page_id: UUID,
    block_id: UUID,
    db: AsyncSession = Depends(get_db),
):
    report = await _load_report(db, project_id, report_id)
    page = _page_from_report(report, page_id)
    block = _block_from_page(page, block_id)
    await db.delete(block)
    await _bump_entity_version(
        db,
        ReportPage,
        page.id,
        detail="页面已被删除，请重新载入",
    )
    await _bump_entity_version(
        db,
        ReportDocument,
        report.id,
        detail="报表已被删除，请重新载入",
    )
    await db.commit()
    return APIResponse.ok(data=await _response_report(db, project_id, report_id))
